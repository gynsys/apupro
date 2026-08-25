import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import re
from difflib import SequenceMatcher
from typing import Optional


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE DIMENSIONES — versión corregida
# ─────────────────────────────────────────────────────────────────────────────

def _norm_unit(raw: str) -> str:
    """Normaliza la unidad de una dimensión."""
    raw = (raw or '').strip().upper()
    return {'': 'IN', '"': 'IN', "''": 'IN', 'PLG': 'IN', 'PULG': 'IN',
            'MM': 'MM', 'CM': 'CM', 'M': 'M'}.get(raw, 'IN')


# Patrón de número (orden importa: mixto > fracción > entero/decimal)
_NUM_PAT = (
    r'(\d+\s+\d+/\d+'   # mixto: "1 1/2"
    r'|\d+/\d+'          # fracción sola: "1/2"  ← debe ir ANTES de \d+
    r'|\d+(?:[.,]\d+)?'  # entero o decimal: "110", "2.5"
    r')'
)
_UNIT_PAT = r'("|\'\'|MM|CM|PLG|PULG|M(?=\b))?'

_RE_DIGUALDAD = re.compile(
    r'\bD\s*=\s*' + _NUM_PAT + r'\s*' + _UNIT_PAT,
    re.IGNORECASE
)
_RE_STANDALONE_MM = re.compile(
    r'(?<![=\d])(\d+(?:[.,]\d+)?)\s*(MM|CM)\b',
    re.IGNORECASE
)
_RE_STANDALONE_IN = re.compile(
    r'(?<![=\d/])(\d+\s+\d+/\d+|\d+/\d+|\d+(?:[.,]\d+)?)\s*(")',
    re.IGNORECASE
)
_RE_X_SEC = re.compile(
    r'\bX\s*' + _NUM_PAT + r'\s*' + _UNIT_PAT,
    re.IGNORECASE
)


def extraer_dimension(texto: str) -> Optional[str]:
    """
    Extrae la dimensión principal de una descripción y la normaliza.
    Detecta:
      - D=110MM, D=1", D=1/2", D=1 1/2"    (con prefijo D=)
      - 20 MM, 110MM                          (MM/CM sin D=)
      - 3", 3/4", 1 1/2"                     (pulgadas sin D=)
    Incluye sección X si existe: D=110MM X 1/2" → "110MMX1/2IN"
    """
    if not texto:
        return None
    t = texto.upper()

    # ── 1. Patrón explícito D= ────────────────────────────────────────────
    m = _RE_DIGUALDAD.search(t)
    if m:
        val = re.sub(r'\s+', '', m.group(1))     # "1 1/2" → "11/2"
        unit = _norm_unit(m.group(2) or '')
        dim = f"{val}{unit}"
        # Buscar dimensión cruzada X después del match
        rest = t[m.end():]
        mx = _RE_X_SEC.search(rest)
        if mx:
            vx = re.sub(r'\s+', '', mx.group(1))
            ux = _norm_unit(mx.group(2) or '')
            return f"{dim}X{vx}{ux}"
        return dim

    # ── 2. Standalone MM/CM ───────────────────────────────────────────────
    m = _RE_STANDALONE_MM.search(t)
    if m:
        val = m.group(1).replace(',', '.')
        unit = m.group(2).upper()
        dim = f"{val}{unit}"
        rest = t[m.end():]
        mx = _RE_X_SEC.search(rest)
        if mx:
            vx = re.sub(r'\s+', '', mx.group(1))
            ux = _norm_unit(mx.group(2) or '')
            return f"{dim}X{vx}{ux}"
        return dim

    # ── 3. Standalone pulgadas ────────────────────────────────────────────
    m = _RE_STANDALONE_IN.search(t)
    if m:
        val = re.sub(r'\s+', '', m.group(1))
        return f"{val}IN"

    return None


def dimensiones_compatibles(d1: Optional[str], d2: Optional[str]) -> bool:
    """
    True si las dimensiones son compatibles para hacer match.
    - Ambas None → compatible (sin dimensión)
    - Una None, otra con valor → compatible (no aplica restricción)
    - Ambas con valor → deben ser iguales
    """
    if d1 is None or d2 is None:
        return True
    return d1 == d2


# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZACIÓN DE TEXTO (sin dimensión — se compara por separado)
# ─────────────────────────────────────────────────────────────────────────────

def normalizar_descripcion(texto: str) -> str:
    if not texto:
        return ""

    t = texto.upper()

    # Quitar TIPO "X" → TIPO X antes del resto
    t = re.sub(r'\bTIPO\s+"([A-Z])"\s*', r'TIPO \1 ', t)

    # Eliminar # 
    t = re.sub(r'#', '', t)

    # Puntos entre letras
    t = re.sub(r'([A-Z])\.([A-Z])', r'\1\2', t)
    t = re.sub(r'([A-Z])\.([A-Z])', r'\1\2', t)

    # Unidades
    t = re.sub(r'\bKGF\.?\b', 'KG', t)
    t = re.sub(r'\bKG\.?\b',  'KG', t)
    t = re.sub(r'\bLTS?\b',   'LITRO', t)
    t = re.sub(r'\bMTS?\b',   'METRO', t)
    t = re.sub(r'\bCMS?\b',   'CENTIMETRO', t)
    t = re.sub(r'\bMMS?\b',   'MILIMETRO', t)
    t = re.sub(r'\bPZS?\b',   'PIEZA', t)
    t = re.sub(r'\bUNDS?\b',  'UNIDAD', t)

    # Separador de miles
    t = re.sub(r'(\d{1,3})\.(\d{3})(?!\d)', r'\1\2', t)

    # ── ELIMINAR TODA DIMENSIÓN del texto normalizado ─────────────────────
    # D= con su valor y unidad
    t = re.sub(r'\bD\s*=\s*[\d\s/.,]+(?:"|\'\'|MM|CM|PLG|PULG|M\b)?', '', t)
    # X dimensión cruzada
    t = re.sub(r'\bX\s*[\d\s/.,]+(?:"|\'\'|MM|CM|PLG|PULG|M\b)?', '', t)
    # Standalone MM/CM
    t = re.sub(r'(?<![=\d])\d+(?:[.,]\d+)?\s*(?:MM|CM)\b', '', t)
    # Standalone pulgadas (números seguidos de ")
    t = re.sub(r'(?<![=\d/])\d+(?:\s+\d+/\d+|\d+/\d+)?\s*"', '', t)

    # Guiones → espacio
    t = re.sub(r'-', ' ', t)

    # Limpiar espacios
    t = re.sub(r'\s+', ' ', t).strip().lower()

    # Eliminar puntuación restante
    t = re.sub(r'[^\w\s]', ' ', t)

    # Stop words — conservar números (discriminan: 3 vs 4 tornillos)
    stop_words = {
        'con', 'sin', 'para', 'de', 'la', 'el', 'los', 'las',
        'un', 'una', 'y', 'o', 'en', 'por'
    }
    words = t.split()
    # Conservar: longitud > 2 O es número (discrimina cantidades)
    words = [w for w in words if w not in stop_words and (len(w) > 2 or w.isdigit())]

    return ' '.join(sorted(words))


def calcular_similitud(t1: str, t2: str) -> float:
    return SequenceMatcher(None, t1, t2).ratio()


def extraer_palabras_clave(texto: str) -> list:
    if not texto:
        return []
    t = re.sub(r'[^\w\s]', ' ', texto.lower())
    stop_words = {
        'con', 'sin', 'para', 'de', 'la', 'el', 'los', 'las',
        'un', 'una', 'y', 'o', 'en', 'tipo', 'por'
    }
    # Conservar palabras de más de 3 chars O números (3, 4, 10...)
    return [w for w in t.split()
            if w not in stop_words and (len(w) > 3 or w.isdigit())]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Conectando a la base de datos...")
    engine = create_engine(
        'postgresql://apupro_user:apupro_password@costbase.net:5440/apupro_db'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Analisis MAT v3"

    header_fill      = PatternFill(start_color="1A6BB5", end_color="1A6BB5", fill_type="solid")
    header_font      = Font(bold=True, color="FFFFFF", size=11)
    exacto_fill      = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    parcial_fill     = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    sin_match_fill   = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    headers = [
        "Codigo MAT", "Descripcion MAT", "Precio MAT", "Dimension MAT",
        "Match", "Codigo Match", "Descripcion Match", "Precio Match", "Dimension Match",
        "Score", "Dif $", "Dif %",
        "Tipo Match", "Propuesta", "Cod Propuesto", "Observaciones"
    ]
    col_widths = [13, 44, 10, 14, 5, 13, 44, 10, 14, 8, 10, 8, 12, 12, 13, 34]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

    with engine.connect() as conn:
        print("Obteniendo MAT...")
        todos_mat = conn.execute(text(
            'SELECT "CodMat", "Descri", "CosMat" FROM cost360_materials '
            "WHERE \"CodMat\" LIKE 'MAT%' ORDER BY \"CodMat\""
        )).fetchall()
        print(f"Total MAT: {len(todos_mat)}")

        print("Obteniendo no-MAT...")
        todos_no_mat = conn.execute(text(
            'SELECT "CodMat", "Descri", "CosMat" FROM cost360_materials '
            "WHERE \"CodMat\" NOT LIKE 'MAT%' ORDER BY \"CodMat\""
        )).fetchall()
        print(f"Total no-MAT: {len(todos_no_mat)}")

        # Índice
        mdict: dict = {}
        kw_idx: dict = {}
        for cod, descri, precio in todos_no_mat:
            norm   = normalizar_descripcion(descri)
            kws    = extraer_palabras_clave(descri)
            dim    = extraer_dimension(descri)
            mdict[cod] = {
                'descri': descri, 'precio': precio,
                'norm': norm, 'kws': kws, 'dim': dim
            }
            for kw in kws:
                kw_idx.setdefault(kw, []).append(cod)

        row_num = 2
        stats = {'exacto': 0, 'parcial': 0, 'sin_match': 0}

        print("Procesando con validacion de dimension...")

        for mat_cod, mat_descri, mat_precio in todos_mat:
            mat_norm = normalizar_descripcion(mat_descri)
            mat_kws  = extraer_palabras_clave(mat_descri)
            mat_dim  = extraer_dimension(mat_descri)

            # ── EXACTO ────────────────────────────────────────────────────
            match_exacto = None
            for cod, d in mdict.items():
                if d['norm'] == mat_norm and dimensiones_compatibles(mat_dim, d['dim']):
                    match_exacto = cod
                    break

            # ── KEYWORDS ──────────────────────────────────────────────────
            match_kw = None
            if not match_exacto and len(mat_kws) >= 2:
                candidatos: set = set()
                for kw in mat_kws:
                    candidatos.update(kw_idx.get(kw, []))

                mejor = 0
                for cod in candidatos:
                    d = mdict[cod]
                    if not dimensiones_compatibles(mat_dim, d['dim']):
                        continue
                    hit = sum(1 for kw in mat_kws if kw in d['kws'])
                    if hit > mejor and hit >= len(mat_kws) * 0.75:
                        match_kw = cod
                        mejor = hit

            # ── FUZZY ─────────────────────────────────────────────────────
            match_fuzzy = None
            best_score = 0.0
            if not match_exacto and not match_kw and len(mat_kws) >= 1:
                candidatos = set()
                for kw in mat_kws:
                    candidatos.update(kw_idx.get(kw, []))
                if not candidatos:
                    import random
                    candidatos = set(random.sample(
                        list(mdict.keys()), min(100, len(mdict))
                    ))
                for cod in candidatos:
                    d = mdict[cod]
                    if not dimensiones_compatibles(mat_dim, d['dim']):
                        continue
                    sc = calcular_similitud(mat_norm, d['norm'])
                    if sc > 0.88 and sc > best_score:
                        match_fuzzy = cod
                        best_score = sc

            # ── Resultado ─────────────────────────────────────────────────
            if match_exacto:
                tipo = "EXACTO"; hit_cod = match_exacto
                fill = exacto_fill; prop = "FUSIONAR"
                score_str = "100%"; obs = "Exacto normalizado"
                stats['exacto'] += 1
            elif match_fuzzy:
                tipo = "FUZZY"; hit_cod = match_fuzzy
                fill = parcial_fill
                prop = "REVISAR" if best_score < 0.92 else "FUSIONAR"
                score_str = f"{best_score:.1%}"; obs = f"Fuzzy {best_score:.1%}"
                stats['parcial'] += 1
            elif match_kw:
                tipo = "KEYWORDS"; hit_cod = match_kw
                fill = parcial_fill; prop = "REVISAR"
                score_str = "KW"; obs = "Keywords match"
                stats['parcial'] += 1
            else:
                tipo = "SIN MATCH"; hit_cod = None
                fill = sin_match_fill; prop = "MANUAL"
                score_str = ""; obs = ""
                stats['sin_match'] += 1

            if hit_cod:
                hd = mdict[hit_cod]
                h_descri = hd['descri']; h_precio = hd['precio']; h_dim = hd['dim']
            else:
                h_descri = h_precio = h_dim = hit_cod = ""

            diff_d = abs(mat_precio - (h_precio or 0)) if h_precio else 0
            diff_p = ((h_precio - mat_precio) / mat_precio * 100
                      if mat_precio and h_precio else 0)

            row_data = [
                mat_cod, mat_descri, round(mat_precio, 2), mat_dim or "",
                "SI" if hit_cod else "NO",
                hit_cod or "", h_descri, round(h_precio, 2) if h_precio else "", h_dim or "",
                score_str,
                round(diff_d, 2) if diff_d else "",
                f"{diff_p:.1f}%" if diff_p else "",
                tipo, prop, hit_cod or "", obs
            ]

            for i, val in enumerate(row_data, 1):
                c = ws.cell(row=row_num, column=i, value=val)
                c.fill = fill; c.border = thin
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                if i in [3, 8, 11] and val and isinstance(val, (int, float)):
                    c.number_format = '#,##0.00'

            row_num += 1
            if row_num % 100 == 0:
                print(f"Procesados {row_num - 1} de {len(todos_mat)}...")

    total = len(todos_mat)
    print("\n" + "=" * 60)
    print("RESUMEN v3 (dimension + numeros como discriminador)")
    print("=" * 60)
    for k, v in stats.items():
        print(f"  {k:<12}: {v:>5}  ({v/total*100:.1f}%)")

    fname = f"analisis_mat_v3_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    print(f"\nArchivo: {fname}")


if __name__ == '__main__':
    main()
