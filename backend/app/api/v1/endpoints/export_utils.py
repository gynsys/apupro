from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
import uuid
import base64
import io
import re
from app.core.logging import logger

# Eliminamos numero_a_letras ya que vendrá del frontend

def style_cell(
    cell: Any,
    bold: bool = False,
    size: int = 11,
    align: str = "left",
    valign: str = "center",
    border: bool = False,
    number_format: Optional[str] = None,
    fill: Optional[Any] = None,
    font_color: str = "000000"
) -> None:
    cell.font = Font(bold=bold, size=size, name="Calibri", color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)
    if border:
        thin = Side(style='thin', color='CBD5E1')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill

def generate_excel_workbook(item, mat_rows, eq_rows, mo_rows, settings=None):
    """Genera Excel con el formato original (apu_formulas.py) usando diccionarios"""
    if settings is None:
        settings = {}
        
    rendimiento = float(item.get("RenPar") or item.get("rendimiento") or item.get("performance") or 1.0)
    cantidad = float(item.get("CanPar") or item.get("cantidad") or item.get("quantity") or 1.0)
    descripcion = item.get("DesPar") or item.get("descripcion") or item.get("description") or ""
    codigo = item.get("cov_par") or item.get("CovPar") or item.get("codigo_covenin") or item.get("CodPar") or item.get("codigo") or item.get("code") or ""
    unidad = item.get("UniPar") or item.get("unidad") or item.get("unit") or ""
    admin_gg = float(settings.get("admin_percent", 15.0))
    imprevisto_ut = float(settings.get("profit_percent", 10.0))
    financiamiento = float(settings.get("financiamiento", 0.0))
    iva = float(settings.get("iva_percent", 16.0))
    otros_imp = float(settings.get("otros_imp", 0.0))
    prestaciones = float(settings.get("fcas_percent", 435.0))

    wb = Workbook()
    ws = wb.active
    ws.title = f"APU_{item.get('CodPar', 'Custom')}"

    # HEADER
    ws.merge_cells("B1:H1")
    ws["B1"] = "ANÁLISIS DE PRECIO UNITARIO"
    style_cell(ws["B1"], bold=True, size=14, align="center")
    
    ws.merge_cells("B2:H2")
    ws["B2"] = ""
    
    # Ajustar ancho de la columna B para que quepa "Descripción"
    ws.column_dimensions['B'].width = 10
    
    ws["B3"] = f"Obra: {settings.get('project_name') or ''}"
    ws["B4"] = f"Contratante: {settings.get('client_name') or ''}"
    ws["E5"] = "Part. No.:"
    ws["F5"] = "1"
    ws["G5"] = "Fecha:"
    ws["H5"] = datetime.now().strftime("%d/%m/%Y")
    ws["B6"] = "Descripción:"
    ws.merge_cells("C6:H7")
    ws["C6"] = item.get("Descri", 'N/A')
    style_cell(ws["C6"], align="left", valign="top")
    ws["G8"] = "Rendimiento:"
    ws["H8"] = rendimiento
    ws["B9"] = "Código:"
    ws["C9"] = codigo
    ws["E9"] = "Unidad:"
    ws["F9"] = item.get("UniPar", "UND")
    ws["G9"] = "Cantidad:"
    ws["H9"] = "1"

    # MATERIALES
    mat_start = 11
    ws.merge_cells(f"B{mat_start}:H{mat_start}")
    ws[f"B{mat_start}"] = "MATERIALES"
    style_cell(ws[f"B{mat_start}"], bold=True, size=12)
    
    headers = ["No.", "Descripción", "Und.", "Cant.", "Desp.", "Precio", "Total"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 2)
        ws[f"{col}{mat_start+1}"] = h
        style_cell(ws[f"{col}{mat_start+1}"], bold=True, border=True)
    
    row = mat_start + 2
    for i, m in enumerate(mat_rows):
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = m.get("Descri") or m.get("descripcion", "")
        ws[f"D{row}"] = m.get("UniMat") or m.get("unidad", "")
        ws[f"E{row}"] = float(m.get("CanIns") or m.get("cantidad") or 0)
        ws[f"F{row}"] = float(m.get("Desper") or m.get("desperdicio") or 0)
        ws[f"G{row}"] = float(m.get("CosMat") or m.get("precio_unitario") or m.get("precio") or 0)
        ws[f"H{row}"] = f"=ROUND((G{row}*E{row})*((F{row}/100)+1),2)"
        style_cell(ws[f"H{row}"], number_format='#,##0.00')
        row += 1
    
    total_mat_row = row
    ws[f"F{total_mat_row}"] = "Total Materiales:"
    first_data = mat_start + 2
    last_data = total_mat_row - 1 if row > mat_start + 2 else mat_start + 2
    ws[f"H{total_mat_row}"] = f"=SUM(H{first_data}:H{last_data})"
    style_cell(ws[f"H{total_mat_row}"], bold=True, number_format='#,##0.00')

    # EQUIPOS
    eq_start = total_mat_row + 2
    ws.merge_cells(f"B{eq_start}:H{eq_start}")
    ws[f"B{eq_start}"] = "EQUIPOS"
    style_cell(ws[f"B{eq_start}"], bold=True, size=12)
    
    headers = ["No.", "Descripción", "", "Cant.", "Cop/Dep", "Precio", "Total"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 2)
        ws[f"{col}{eq_start+1}"] = h
        style_cell(ws[f"{col}{eq_start+1}"], bold=True, border=True)
    
    row = eq_start + 2
    for i, e in enumerate(eq_rows):
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = e.get("Descri") or e.get("descripcion", "")
        ws[f"E{row}"] = float(e.get("CanIns") or e.get("cantidad") or 0)
        ws[f"F{row}"] = float(e.get("Deprec") or e.get("depreciacion") or 1)
        ws[f"G{row}"] = float(e.get("CosDia") or e.get("precio_unitario") or e.get("precio") or 0)
        ws[f"H{row}"] = f"=ROUND((G{row}*E{row})*(F{row}),2)"
        style_cell(ws[f"H{row}"], number_format='#,##0.00')
        row += 1
    
    total_eq_row = row
    ws[f"F{total_eq_row}"] = "Total Equipos:"
    first_data = eq_start + 2
    last_data = total_eq_row - 1 if row > eq_start + 2 else eq_start + 2
    ws[f"H{total_eq_row}"] = f"=SUM(H{first_data}:H{last_data})"
    style_cell(ws[f"H{total_eq_row}"], bold=True, number_format='#,##0.00')
    
    cuo_row = total_eq_row + 1
    ws[f"E{cuo_row}"] = "Costo Unitarios Equipos:"
    ws[f"H{cuo_row}"] = f"=ROUND(H{total_eq_row}/H8,2)"
    style_cell(ws[f"H{cuo_row}"], bold=True, number_format='#,##0.00')

    # MANO DE OBRA
    mo_start = cuo_row + 2
    ws.merge_cells(f"B{mo_start}:H{mo_start}")
    ws[f"B{mo_start}"] = "MANO DE OBRA"
    style_cell(ws[f"B{mo_start}"], bold=True, size=12)
    
    headers = ["No.", "Descripción", "Cant.", "Jornal", "Bono", "Total Jornal", "Total Bono"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 2)
        ws[f"{col}{mo_start+1}"] = h
        style_cell(ws[f"{col}{mo_start+1}"], bold=True, border=True)
    
    row = mo_start + 2
    for i, l in enumerate(mo_rows):
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = l.get("Descri") or l.get("descripcion", "")
        ws[f"D{row}"] = float(l.get("CanIns") or l.get("cantidad") or 0)
        ws[f"E{row}"] = float(l.get("Jornal") or l.get("jornal") or 0)
        ws[f"F{row}"] = float(l.get("Bono") or l.get("bono") or 0)
        ws[f"G{row}"] = f"=ROUND((D{row}*E{row}),2)"
        ws[f"H{row}"] = f"=ROUND((D{row}*F{row}),2)"
        style_cell(ws[f"G{row}"], number_format='#,##0.00')
        style_cell(ws[f"H{row}"], number_format='#,##0.00')
        row += 1
    
    sub_row = row
    first_data = mo_start + 2
    last_data = sub_row - 1 if row > mo_start + 2 else mo_start + 2
    ws[f"D{sub_row}"] = "SubTotal Mano de Obra:"
    ws[f"G{sub_row}"] = f"=SUM(G{first_data}:G{last_data})"
    ws[f"H{sub_row}"] = f"=SUM(H{first_data}:H{last_data})"
    style_cell(ws[f"G{sub_row}"], bold=True, number_format='#,##0.00')
    style_cell(ws[f"H{sub_row}"], bold=True, number_format='#,##0.00')
    
    ps_row = sub_row + 1
    ws[f"C{ps_row}"] = prestaciones
    style_cell(ws[f"C{ps_row}"], number_format='#,##0.00', align="right")
    ws[f"D{ps_row}"] = "Prestaciones Sociales:"
    ws[f"G{ps_row}"] = f"=ROUND((C{ps_row}/100)*G{sub_row},2)"
    ws[f"H{ps_row}"] = 0
    style_cell(ws[f"G{ps_row}"], number_format='#,##0.00')
    
    tg_row = ps_row + 1
    ws[f"D{tg_row}"] = "Total General Mano de Obra:"
    ws[f"H{tg_row}"] = f"=G{ps_row}+H{ps_row}+G{sub_row}+H{sub_row}"
    style_cell(ws[f"H{tg_row}"], bold=True, number_format='#,##0.00')
    
    cuo_mo_row = tg_row + 1
    ws[f"D{cuo_mo_row}"] = "Costo Unitario de Mano de Obra:"
    ws[f"H{cuo_mo_row}"] = f"=ROUND(H{tg_row}/H8,2)"
    style_cell(ws[f"H{cuo_mo_row}"], bold=True, number_format='#,##0.00')

    # RESUMEN
    resumen_start = cuo_mo_row + 2
    cd_row = resumen_start + 1
    ws[f"E{cd_row}"] = "COSTO DIRECTO SUBTOTAL A:"
    ws[f"H{cd_row}"] = f"=ROUND(H{total_mat_row}+H{cuo_row}+H{cuo_mo_row},2)"
    style_cell(ws[f"H{cd_row}"], bold=True, number_format='#,##0.00')
    
    ad_row = cd_row + 1
    ws[f"C{ad_row}"] = admin_gg
    style_cell(ws[f"C{ad_row}"], number_format='#,##0.00', align="right")
    ws[f"D{ad_row}"] = "Administración y Gastos Generales:"
    ws[f"H{ad_row}"] = f"=ROUND((H{cd_row}*C{ad_row})/100,2)"
    style_cell(ws[f"H{ad_row}"], number_format='#,##0.00')
    
    sb_row = ad_row + 1
    ws[f"D{sb_row}"] = "SUBTOTAL B:"
    ws[f"H{sb_row}"] = f"=H{cd_row}+H{ad_row}"
    style_cell(ws[f"H{sb_row}"], bold=True, number_format='#,##0.00')
    
    iu_row = sb_row + 1
    ws[f"E{iu_row}"] = imprevisto_ut
    style_cell(ws[f"E{iu_row}"], number_format='#,##0.00', align="right")
    ws[f"F{iu_row}"] = "Imprevisto Utilidad:"
    ws[f"H{iu_row}"] = f"=ROUND((H{sb_row}*E{iu_row})/100,2)"
    style_cell(ws[f"H{iu_row}"], number_format='#,##0.00')
    
    sc_row = iu_row + 1
    ws[f"D{sc_row}"] = "SUBTOTAL C:"
    ws[f"H{sc_row}"] = f"=H{sb_row}+H{iu_row}"
    style_cell(ws[f"H{sc_row}"], bold=True, number_format='#,##0.00')
    
    fin_row = sc_row + 1
    ws[f"E{fin_row}"] = financiamiento
    style_cell(ws[f"E{fin_row}"], number_format='#,##0.00', align="right")
    ws[f"F{fin_row}"] = "Financiamiento:"
    ws[f"H{fin_row}"] = f"=ROUND((H{sc_row}*E{fin_row})/100,2)"
    style_cell(ws[f"H{fin_row}"], number_format='#,##0.00')
    
    ps_row2 = fin_row + 1
    ws[f"D{ps_row2}"] = "PRECIO UNITARIO SIN IMPUESTO:"
    ws[f"H{ps_row2}"] = f"=H{sc_row}+H{fin_row}"
    style_cell(ws[f"H{ps_row2}"], bold=True, number_format='#,##0.00')
    
    iva_row = ps_row2 + 1
    ws[f"E{iva_row}"] = iva
    style_cell(ws[f"E{iva_row}"], number_format='#,##0.00', align="right")
    ws[f"F{iva_row}"] = "Impuesto (I.V.A.):"
    ws[f"H{iva_row}"] = f"=ROUND((H{ps_row2}*E{iva_row})/100,2)"
    style_cell(ws[f"H{iva_row}"], number_format='#,##0.00')
    
    oi_row = iva_row + 1
    ws[f"E{oi_row}"] = otros_imp
    style_cell(ws[f"E{oi_row}"], number_format='#,##0.00', align="right")
    ws[f"F{oi_row}"] = "Otros Impuestos:"
    ws[f"H{oi_row}"] = f"=ROUND((H{ps_row2}*E{oi_row})/100,2)"
    style_cell(ws[f"H{oi_row}"], number_format='#,##0.00')
    
    pf_row = oi_row + 2
    ws[f"D{pf_row}"] = "PRECIO UNITARIO (Bs.F.):"
    ws[f"H{pf_row}"] = f"=H{ps_row2}+H{iva_row}+H{oi_row}"
    style_cell(ws[f"H{pf_row}"], bold=True, number_format='#,##0.00')
    
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18

    # Convertir el precio total a letras en B{cd_row}:C{pf_row-1}
    total_mat = round(sum( round((float(m.get("CosMat") or m.get("precio_unitario") or m.get("precio") or 0) * float(m.get("CanIns") or m.get("cantidad") or 0)) * (float(m.get("Desper") or m.get("desperdicio") or 0)/100.0 + 1), 2) for m in mat_rows ), 2)
    total_eq = round(sum( round((float(e.get("CosDia") or e.get("precio_unitario") or e.get("precio") or 0) * float(e.get("CanIns") or e.get("cantidad") or 0)) * float(e.get("Deprec") or e.get("depreciacion") or 1), 2) for e in eq_rows ), 2)
    total_mo_jornal = sum( round(float(l.get("CanIns") or l.get("cantidad") or 0) * float(l.get("Jornal") or l.get("jornal") or 0), 2) for l in mo_rows )
    total_mo_bono = sum( round(float(l.get("CanIns") or l.get("cantidad") or 0) * float(l.get("Bono") or l.get("bono") or 0), 2) for l in mo_rows )
    mo_ps = round((prestaciones/100.0) * total_mo_jornal, 2)
    total_mo = total_mo_jornal + total_mo_bono + mo_ps
    
    cuo_eq = round(total_eq / rendimiento, 2)
    cuo_mo = round(total_mo / rendimiento, 2)
    
    cd_val = round(total_mat + cuo_eq + cuo_mo, 2)
    ad_val = round((cd_val * admin_gg)/100, 2)
    sb_val = cd_val + ad_val
    iu_val = round((sb_val * imprevisto_ut)/100, 2)
    sc_val = sb_val + iu_val
    fin_val = round((sc_val * financiamiento)/100, 2)
    ps_val = sc_val + fin_val
    iva_val = round((ps_val * iva)/100, 2)
    oi_val = round((ps_val * otros_imp)/100, 2)
    precio_final = ps_val + iva_val + oi_val

    # El texto ya viene calculado desde el frontend en settings
    son_letras = settings.get("son_letras", "SON: ( NO DISPONIBLE )")

    ws.merge_cells(f"B{iva_row}:C{pf_row-1}")
    ws[f"B{iva_row}"] = son_letras
    style_cell(ws[f"B{iva_row}"], align="center", valign="bottom", bold=True)

    # Alinear al centro toda la columna H (valores y totales)
    for r in range(1, pf_row + 1):
        cell = ws[f"H{r}"]
        if cell.value is not None:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Guardar
    temp_dir = Path("temp")
    temp_dir.mkdir(exist_ok=True)
    filename = f"APU_{item.get('CovPar') or item.get('CodPar', 'Custom')}_{uuid.uuid4().hex[:6]}.xlsx"
    file_path = temp_dir / filename
    wb.save(file_path)

    return file_path, filename


def generate_budget_excel_workbook(
    budget: Dict[str, Any],
    items: List[Dict[str, Any]],
    settings: Optional[Dict[str, Any]] = None
) -> Tuple[Path, str]:
    """Genera un archivo Excel (.xlsx) completo para el presupuesto, respetando configuraciones, capítulos, logo y fórmulas."""
    if not isinstance(budget, dict):
        raise ValueError("budget debe ser un diccionario válido")
    if not isinstance(items, list):
        raise ValueError("items debe ser una lista válida")

    if settings is None:
        settings = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws.views.sheetView[0].showGridLines = True

    # 1. Definir anchos de columna
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8    # # / N°
    ws.column_dimensions['C'].width = 18   # Código
    ws.column_dimensions['D'].width = 60   # Descripción
    ws.column_dimensions['E'].width = 10   # Unidad
    ws.column_dimensions['F'].width = 14   # Cantidad
    ws.column_dimensions['G'].width = 18   # P.U.
    ws.column_dimensions['H'].width = 20   # Total

    # 2. LOGO (si está activo y presente en base64)
    logo_b64 = budget.get("logo_base64") or settings.get("logo_base64")
    has_logo = False
    if logo_b64:
        try:
            raw_b64 = str(logo_b64)
            if "," in raw_b64:
                raw_b64 = raw_b64.split(",", 1)[1]
            img_bytes = base64.b64decode(raw_b64)
            pil_img = PILImage.open(io.BytesIO(img_bytes))
            orig_w, orig_h = pil_img.size
            if orig_h > 0 and orig_w > 0:
                max_h = 52
                aspect = orig_w / orig_h
                new_h = min(orig_h, max_h)
                new_w = int(new_h * aspect)

                output_buf = io.BytesIO()
                pil_img.convert("RGBA").save(output_buf, format="PNG")
                output_buf.seek(0)

                xl_img = OpenpyxlImage(output_buf)
                xl_img.width = new_w
                xl_img.height = new_h
                ws.add_image(xl_img, 'B2')
                has_logo = True
        except Exception as e:
            logger.error(f"Error procesando logo para Excel: {e}", exc_info=True)

    # 3. BLOQUE DE DATOS DE CABECERA
    info_col = "D" if has_logo else "B"
    header_row = 2

    obra = (budget.get("obra") or budget.get("project_name") or budget.get("name") or "").strip()
    if obra:
        ws[f"{info_col}{header_row}"] = f"Obra: {obra}"
        ws[f"{info_col}{header_row}"].font = Font(bold=True, size=11, name="Calibri")
        header_row += 1

    ubicacion = (budget.get("ubicacion") or settings.get("ubicacion") or "").strip()
    if ubicacion:
        ws[f"{info_col}{header_row}"] = f"Ubicación: {ubicacion}"
        ws[f"{info_col}{header_row}"].font = Font(size=10, name="Calibri", color="475569")
        header_row += 1

    contratante = (budget.get("contratante") or budget.get("client_name") or "").strip()
    if contratante:
        ws[f"{info_col}{header_row}"] = f"Contratante: {contratante}"
        ws[f"{info_col}{header_row}"].font = Font(size=10, name="Calibri", color="475569")
        header_row += 1

    company_rif = (budget.get("company_rif") or "").strip()
    if company_rif:
        ws[f"{info_col}{header_row}"] = f"RIF: {company_rif}"
        ws[f"{info_col}{header_row}"].font = Font(size=10, name="Calibri", color="475569")
        header_row += 1

    ws[f"{info_col}{header_row}"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws[f"{info_col}{header_row}"].font = Font(size=9, name="Calibri", color="64748B")
    header_row += 1

    if has_logo and header_row < 6:
        header_row = 6

    # 4. TÍTULO CENTRADO
    title_row = header_row + 1
    title_text = (budget.get("title") or "PRESUPUESTO").strip().upper()
    ws.merge_cells(f"B{title_row}:H{title_row}")
    ws[f"B{title_row}"] = title_text
    ws[f"B{title_row}"].font = Font(bold=True, size=14, name="Calibri", color="0F172A")
    ws[f"B{title_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 26

    # Moneda
    raw_curr = str(budget.get("currency") or "USD").strip().upper()
    currency_label = "Bs." if ("BS" in raw_curr) else "USD"

    # 5. ENCABEZADOS DE LA TABLA
    th_row = title_row + 2
    headers = [
        ("B", "#", "center"),
        ("C", "CÓDIGO", "center"),
        ("D", "DESCRIPCIÓN", "left"),
        ("E", "UNIDAD", "center"),
        ("F", "CANTIDAD", "right"),
        ("G", f"P.U. ({currency_label})", "right"),
        ("H", f"TOTAL ({currency_label})", "right")
    ]
    th_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    for col, text, align in headers:
        cell = ws[f"{col}{th_row}"]
        cell.value = text
        style_cell(cell, bold=True, size=11, align=align, border=True, fill=th_fill, font_color="1E293B")
    ws.row_dimensions[th_row].height = 24

    # 6. FILAS DE DATOS (PARTIDAS Y CAPÍTULOS)
    curr_r = th_row + 1
    first_data_r = curr_r
    item_num = 1
    chap_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for it in items:
        is_chapter = bool(it.get("is_chapter"))
        desc = str(it.get("description") or "").strip()

        if is_chapter:
            ws.merge_cells(f"B{curr_r}:H{curr_r}")
            cell = ws[f"B{curr_r}"]
            cell.value = desc.upper()
            style_cell(cell, bold=True, size=11, align="left", border=True, fill=chap_fill, font_color="0F172A")
            for c in ["C", "D", "E", "F", "G", "H"]:
                style_cell(ws[f"{c}{curr_r}"], border=True, fill=chap_fill)
            ws.row_dimensions[curr_r].height = 22
        else:
            style_cell(ws[f"B{curr_r}"], align="center", border=True)
            ws[f"B{curr_r}"] = item_num

            style_cell(ws[f"C{curr_r}"], align="center", border=True)
            ws[f"C{curr_r}"] = str(it.get("cov_par") or it.get("cod_par") or "")

            style_cell(ws[f"D{curr_r}"], align="left", border=True)
            ws[f"D{curr_r}"] = desc

            style_cell(ws[f"E{curr_r}"], align="center", border=True)
            ws[f"E{curr_r}"] = str(it.get("unit") or "")

            qty = float(it.get("quantity") or 0.0)
            style_cell(ws[f"F{curr_r}"], align="right", border=True, number_format="#,##0.00")
            ws[f"F{curr_r}"] = qty

            pu = float(it.get("pu") or it.get("unit_price") or 0.0)
            style_cell(ws[f"G{curr_r}"], align="right", border=True, number_format="#,##0.00")
            ws[f"G{curr_r}"] = pu

            style_cell(ws[f"H{curr_r}"], align="right", border=True, number_format="#,##0.00")
            ws[f"H{curr_r}"] = f"=ROUND(F{curr_r}*G{curr_r}, 2)"

            item_num += 1
            ws.row_dimensions[curr_r].height = 20

        curr_r += 1

    last_data_r = max(first_data_r, curr_r - 1)

    # 7. TOTALES AL PIE
    tot_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    highlight_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

    # Subtotal
    sub_r = curr_r
    ws.merge_cells(f"B{sub_r}:G{sub_r}")
    ws[f"B{sub_r}"] = f"Subtotal ({currency_label}):"
    style_cell(ws[f"B{sub_r}"], bold=True, size=11, align="right", border=True, fill=tot_fill)
    for c in ["C", "D", "E", "F", "G"]:
        style_cell(ws[f"{c}{sub_r}"], border=True, fill=tot_fill)
    style_cell(ws[f"H{sub_r}"], bold=True, size=11, align="right", border=True, number_format="#,##0.00", fill=tot_fill)
    ws[f"H{sub_r}"] = f"=SUM(H{first_data_r}:H{last_data_r})"
    ws.row_dimensions[sub_r].height = 22

    # IVA
    iva_r = sub_r + 1
    iva_pct = float(budget.get("iva_percent") if budget.get("iva_percent") is not None else 16.0)
    iva_display = f"{int(iva_pct)}%" if iva_pct.is_integer() else f"{iva_pct}%"
    ws.merge_cells(f"B{iva_r}:G{iva_r}")
    ws[f"B{iva_r}"] = f"IVA {iva_display} ({currency_label}):"
    style_cell(ws[f"B{iva_r}"], bold=True, size=11, align="right", border=True, fill=tot_fill)
    for c in ["C", "D", "E", "F", "G"]:
        style_cell(ws[f"{c}{iva_r}"], border=True, fill=tot_fill)
    style_cell(ws[f"H{iva_r}"], bold=True, size=11, align="right", border=True, number_format="#,##0.00", fill=tot_fill)
    ws[f"H{iva_r}"] = f"=ROUND(H{sub_r}*({iva_pct}/100), 2)"
    ws.row_dimensions[iva_r].height = 22

    # Total Presupuesto
    total_r = iva_r + 1
    ws.merge_cells(f"B{total_r}:G{total_r}")
    ws[f"B{total_r}"] = f"Total Presupuesto ({currency_label}):"
    style_cell(ws[f"B{total_r}"], bold=True, size=12, align="right", border=True, fill=highlight_fill, font_color="0F172A")
    for c in ["C", "D", "E", "F", "G"]:
        style_cell(ws[f"{c}{total_r}"], border=True, fill=highlight_fill)
    style_cell(ws[f"H{total_r}"], bold=True, size=12, align="right", border=True, number_format="#,##0.00", fill=highlight_fill, font_color="0F172A")
    ws[f"H{total_r}"] = f"=H{sub_r}+H{iva_r}"

    double_bottom = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='double', color='0F172A')
    )
    for c in ["B", "C", "D", "E", "F", "G", "H"]:
        ws[f"{c}{total_r}"].border = double_bottom
    ws.row_dimensions[total_r].height = 26

    # 8. GUARDAR ARCHIVO TEMPORAL
    temp_dir = Path("temp")
    temp_dir.mkdir(exist_ok=True)
    raw_name = obra or budget.get("name") or "Presupuesto"
    safe_slug = re.sub(r'[^a-zA-Z0-9_\-]', '_', raw_name).strip('_')[:40] or "Presupuesto"
    filename = f"Presupuesto_{safe_slug}_{uuid.uuid4().hex[:6]}.xlsx"
    file_path = temp_dir / filename
    wb.save(file_path)

    return file_path, filename
