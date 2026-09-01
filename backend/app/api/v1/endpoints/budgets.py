from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List
import os
import uuid
import shutil
from pathlib import Path
from fastapi.responses import FileResponse
from app.db.base import get_db
from app.db.models.budget import Budget, BudgetItem, BudgetAPUMaterial as DBMaterial, BudgetAPUEquipment as DBEquipment, BudgetAPULabor as DBLabor
from app.db.models.cost360 import CostItem, CostAPUMaterial, CostAPUEquipment, CostAPULabor
from app.db.models.backup_logs import BackupLog
from app.schemas.budget import (
    Budget as BudgetSchema, BudgetCreate, BudgetUpdate, BudgetSummary,
    BudgetItemCreate, BudgetItem as BudgetItemSchema, BudgetItemUpdate,
    BudgetAPUMaterialBase, BudgetAPUMaterial, BudgetAPUMaterialUpdate,
    BudgetAPUEquipmentBase, BudgetAPUEquipment, BudgetAPUEquipmentUpdate,
    BudgetAPULaborBase, BudgetAPULabor, BudgetAPULaborUpdate
)
from app.api.v1.endpoints.arko import get_current_arko_admin
from app.services.encryption_service import encryption_service
from app.services.email import send_backup_email
from app.middleware.plan_limits import check_budget_limit, check_items_limit, check_ai_access

def log_backup_action(db: Session, user_id: str, user_email: str, budget_id: str, budget_name: str, action: str, status: str, error_message: str = None, ip_address: str = None):
    """Registra acciones de backup para auditoría"""
    log_entry = BackupLog(
        user_id=user_id,
        user_email=user_email,
        budget_id=budget_id,
        budget_name=budget_name,
        action=action,
        status=status,
        error_message=error_message,
        ip_address=ip_address
    )
    db.add(log_entry)
    db.commit()

router = APIRouter()

@router.post("/", response_model=BudgetSchema, status_code=status.HTTP_201_CREATED)
def create_budget(budget_in: BudgetCreate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    # Verificar límite de presupuestos
    check_budget_limit(current_user)

    budget_data = budget_in.model_dump()
    # Asignar automáticamente el user_id del usuario autenticado
    budget_data["user_id"] = str(current_user.id)

    # Aplicar costos_config del usuario como defaults si no se enviaron valores explícitos
    from app.api.v1.endpoints.arko import _get_costos_config
    costos = _get_costos_config(current_user)
    if budget_data.get("profit_percent") is None:
        budget_data["profit_percent"] = costos.porcentajeUtilidad
    if budget_data.get("admin_percent") is None:
        budget_data["admin_percent"] = costos.porcentajeAdministracion
    if budget_data.get("iva_percent") is None:
        budget_data["iva_percent"] = costos.iva
    if budget_data.get("fcas_percent") is None:
        budget_data["fcas_percent"] = costos.fcas

    db_budget = Budget(**budget_data)
    db.add(db_budget)
    db.commit()
    db.refresh(db_budget)
    return db_budget


@router.get("/", response_model=List[BudgetSummary])
def get_budgets(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    # Solo mostrar presupuestos del usuario autenticado
    budgets = db.query(Budget).filter(Budget.user_id == str(current_user.id)).offset(skip).limit(limit).all()
    return budgets

@router.get("/{budget_id}", response_model=BudgetSchema)
def get_budget(budget_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    return budget

@router.put("/{budget_id}", response_model=BudgetSchema)
def update_budget(budget_id: str, budget_in: BudgetUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")

    update_data = budget_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(budget, field, value)

    db.commit()
    db.refresh(budget)
    return budget

@router.delete("/{budget_id}", status_code=204)
def delete_budget(
    budget_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_arko_admin),
) -> None:
    """Elimina un presupuesto y todos sus items (cascade)."""
    budget = db.query(Budget).filter(
        Budget.id == budget_id,
        Budget.user_id == str(current_user.id),
    ).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    db.delete(budget)
    db.commit()

@router.post("/{budget_id}/upload-logo")
async def upload_budget_logo(budget_id: str, logo: UploadFile = File(...), db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    """Sube el logo de la empresa para un presupuesto específico"""
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")

    # Crear directorio para logos si no existe
    upload_dir = Path("public/company_logos")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Generar nombre único para el archivo
    file_extension = logo.filename.split('.')[-1].lower()
    if file_extension not in ['png', 'jpg', 'jpeg']:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PNG, JPG o JPEG")
    
    unique_filename = f"{budget.id}_{uuid.uuid4().hex[:8]}.{file_extension}"
    file_path = upload_dir / unique_filename
    
    # Guardar el archivo
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(logo.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error guardando el archivo: {str(e)}")
    
    # Actualizar el presupuesto con la URL del logo
    logo_url = f"/company_logos/{unique_filename}"
    budget.company_logo = logo_url
    db.commit()
    
    return {"logo_url": logo_url}

@router.post("/{budget_id}/items", response_model=BudgetItemSchema)
def add_item_to_budget(budget_id: str, item_in: BudgetItemCreate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")

    # Verificar límite de partidas
    check_items_limit(current_user, budget_id)
        
    target_order = item_in.order
    if target_order <= 0:
        max_order = db.query(func.max(BudgetItem.order)).filter(BudgetItem.budget_id == budget_id).scalar() or 0
        target_order = max_order + 1
    else:
        db.query(BudgetItem).filter(
            BudgetItem.budget_id == budget_id,
            BudgetItem.order >= target_order
        ).update({BudgetItem.order: BudgetItem.order + 1})
    
    # 1. Crear el BudgetItem (excluir listas de insumos — no son columnas del modelo ORM)
    item_data = item_in.model_dump(exclude={"materials", "equipments", "labors"})
    item_data["order"] = target_order
    db_item = BudgetItem(**item_data, budget_id=budget_id)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    
    # If it is a chapter, skip APU copying
    if item_in.is_chapter:
        return db_item
    
    # 2. Guardar el APU en BudgetAPU
    # PRIORIDAD: Si el frontend envió insumos pre-calculados (con factores de inflación aplicados),
    # se usan directamente. Si no, se copia desde la base maestra de Cost360 (comportamiento original).
    if item_in.materials is not None or item_in.equipments is not None or item_in.labors is not None:
        # --- Ruta A: Frontend proveyó los insumos (base personalizada con factores aplicados) ---
        for mat in (item_in.materials or []):
            db_mat = DBMaterial(
                budget_item_id=db_item.id,
                codigo=mat.codigo,
                descripcion=mat.descripcion,
                unidad=mat.unidad,
                precio_unitario=mat.precio_unitario,
                cantidad=mat.cantidad,
                desperdicio=mat.desperdicio or 0.0
            )
            db.add(db_mat)

        for eq in (item_in.equipments or []):
            db_eq = DBEquipment(
                budget_item_id=db_item.id,
                codigo=eq.codigo,
                descripcion=eq.descripcion,
                unidad=eq.unidad,
                precio_unitario=eq.precio_unitario,
                cantidad=eq.cantidad,
                depreciacion=eq.depreciacion or 1.0
            )
            db.add(db_eq)

        for lab in (item_in.labors or []):
            db_lab = DBLabor(
                budget_item_id=db_item.id,
                codigo=lab.codigo,
                descripcion=lab.descripcion,
                jornal=lab.jornal,
                bono=lab.bono,
                cantidad=lab.cantidad
            )
            db.add(db_lab)

    else:
        # --- Ruta B: Fallback — Copiar desde la base maestra de Cost360 sin factores ---
        cost_item = db.query(CostItem).filter(CostItem.CodPar == item_in.cod_par).first()
        if cost_item:
            for mat in cost_item.apu_materials:
                db_mat = DBMaterial(
                    budget_item_id=db_item.id,
                    codigo=mat.CodIns,
                    descripcion=mat.material.Descri if mat.material else "",
                    unidad=mat.material.UniMat if mat.material else "",
                    precio_unitario=(mat.material.CosMat if (mat.material and mat.material.CosMat is not None) else 0.0),
                    cantidad=mat.CanIns or 0.0,
                    desperdicio=mat.Desper or 0.0
                )
                db.add(db_mat)
                
            for eq in cost_item.apu_equipments:
                precio_diario_depreciado = eq.equipment.CosDia if (eq.equipment and eq.equipment.CosDia is not None) else 0.0
                depreciacion = eq.Deprec if (eq.Deprec is not None and eq.Deprec > 0) else 1.0
                precio_adquisicion = precio_diario_depreciado / depreciacion if depreciacion > 0 else precio_diario_depreciado
                
                db_eq = DBEquipment(
                    budget_item_id=db_item.id,
                    codigo=eq.CodIns,
                    descripcion=eq.equipment.Descri if eq.equipment else "",
                    unidad="Día",
                    precio_unitario=precio_adquisicion,
                    cantidad=eq.CanIns or 0.0,
                    depreciacion=eq.Deprec or 1.0
                )
                db.add(db_eq)
                
            for lab in cost_item.apu_labors:
                db_lab = DBLabor(
                    budget_item_id=db_item.id,
                    codigo=lab.CodIns,
                    descripcion=lab.labor.Descri if lab.labor else "",
                    jornal=(lab.labor.Jornal if (lab.labor and lab.labor.Jornal is not None) else 0.0),
                    bono=(lab.labor.Bono if (lab.labor and lab.labor.Bono is not None) else 0.0),
                    cantidad=lab.CanIns or 0.0
                )
                db.add(db_lab)
        
    db.commit()
    db.refresh(db_item)
        
    return db_item

@router.delete("/{budget_id}/items/{item_id}")
def delete_item_from_budget(budget_id: str, item_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    item = db.query(BudgetItem).filter(BudgetItem.id == item_id, BudgetItem.budget_id == budget_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"ok": True}

@router.post("/{budget_id}/items/reorder")
def reorder_budget_items(budget_id: str, item_ids: List[str], db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
        
    for index, item_id in enumerate(item_ids):
        db.query(BudgetItem).filter(
            BudgetItem.id == item_id, 
            BudgetItem.budget_id == budget_id
        ).update({BudgetItem.order: index + 1})
        
    db.commit()
    return {"ok": True}

@router.post("/{budget_id}/duplicate", response_model=BudgetSchema)
def duplicate_budget(budget_id: str, new_name: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    
    # 1. Duplicate Budget
    new_budget = Budget(
        user_id=str(current_user.id),  # Asignar al mismo usuario
        name=new_name,
        currency=budget.currency,
        exchange_rate=budget.exchange_rate,
        material_inflation=budget.material_inflation,
        equipment_inflation=budget.equipment_inflation,
        labor_inflation=budget.labor_inflation,
        labor_bonus=budget.labor_bonus,
        admin_percent=budget.admin_percent,
        profit_percent=budget.profit_percent,
        fcas_percent=budget.fcas_percent,
        iva_percent=budget.iva_percent,
        project_name=budget.project_name,
        company_name=budget.company_name,
        company_rif=budget.company_rif,
        client_name=budget.client_name
    )
    db.add(new_budget)
    db.commit()
    db.refresh(new_budget)
    
    # 2. Duplicate Items
    for item in budget.items:
        new_item = BudgetItem(
            budget_id=new_budget.id,
            cod_par=item.cod_par,
            cov_par=item.cov_par,
            description=item.description,
            unit=item.unit,
            quantity=item.quantity,
            performance=item.performance,
            order=item.order,
            is_chapter=item.is_chapter
        )
        db.add(new_item)
        db.commit()
        db.refresh(new_item)
        
        # 3. Duplicate APU components for this item
        for mat in item.materials:
            db.add(DBMaterial(
                budget_item_id=new_item.id,
                codigo=mat.codigo,
                descripcion=mat.descripcion,
                unidad=mat.unidad,
                precio_unitario=mat.precio_unitario,
                cantidad=mat.cantidad,
                desperdicio=mat.desperdicio
            ))
        for eq in item.equipments:
            db.add(DBEquipment(
                budget_item_id=new_item.id,
                codigo=eq.codigo,
                descripcion=eq.descripcion,
                unidad=eq.unidad,
                precio_unitario=eq.precio_unitario,
                cantidad=eq.cantidad,
                depreciacion=eq.depreciacion
            ))
        for lab in item.labors:
            db.add(DBLabor(
                budget_item_id=new_item.id,
                codigo=lab.codigo,
                descripcion=lab.descripcion,
                jornal=lab.jornal,
                bono=lab.bono,
                cantidad=lab.cantidad
            ))
            
    db.commit()
    db.refresh(new_budget)
    return new_budget

@router.put("/{budget_id}/items/{item_id}", response_model=BudgetItemSchema)
def update_item_in_budget(budget_id: str, item_id: str, item_in: BudgetItemUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    db_budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not db_budget:
        raise HTTPException(status_code=404, detail="Budget not found")
        
    db_item = db.query(BudgetItem).filter(BudgetItem.id == item_id, BudgetItem.budget_id == budget_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found in budget")
        
    for key, value in item_in.dict(exclude_unset=True).items():
        setattr(db_item, key, value)
        
    db.commit()
    db.refresh(db_item)
    return db_item

@router.post("/{budget_id}/items/{item_id}/materials", response_model=BudgetAPUMaterial)
def add_material_to_item(budget_id: str, item_id: str, material_in: BudgetAPUMaterialBase, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    db_item = db.query(BudgetItem).filter(BudgetItem.id == item_id, BudgetItem.budget_id == budget_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
        
    db_material = DBMaterial(
        budget_item_id=item_id,
        **material_in.dict()
    )
    db.add(db_material)
    db.commit()
    db.refresh(db_material)
    return db_material

@router.post("/{budget_id}/items/{item_id}/equipments", response_model=BudgetAPUEquipment)
def add_equipment_to_item(budget_id: str, item_id: str, equipment_in: BudgetAPUEquipmentBase, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    db_item = db.query(BudgetItem).filter(BudgetItem.id == item_id, BudgetItem.budget_id == budget_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
        
    db_equipment = DBEquipment(
        budget_item_id=item_id,
        **equipment_in.dict()
    )
    db.add(db_equipment)
    db.commit()
    db.refresh(db_equipment)
    return db_equipment

@router.post("/{budget_id}/items/{item_id}/labors", response_model=BudgetAPULabor)
def add_labor_to_item(budget_id: str, item_id: str, labor_in: BudgetAPULaborBase, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    db_item = db.query(BudgetItem).filter(BudgetItem.id == item_id, BudgetItem.budget_id == budget_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
        
    db_labor = DBLabor(
        budget_item_id=item_id,
        **labor_in.dict()
    )
    db.add(db_labor)
    db.commit()
    db.refresh(db_labor)
    return db_labor

@router.put("/{budget_id}/items/{item_id}/materials/{component_id}", response_model=BudgetAPUMaterial)
def update_material_in_item(budget_id: str, item_id: str, component_id: str, comp_in: BudgetAPUMaterialUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    comp = db.query(DBMaterial).filter(DBMaterial.id == component_id, DBMaterial.budget_item_id == item_id).first()
    if not comp: raise HTTPException(status_code=404, detail="Material not found")
    for key, value in comp_in.dict(exclude_unset=True).items():
        setattr(comp, key, value)
    db.commit()
    db.refresh(comp)
    return comp

@router.put("/{budget_id}/items/{item_id}/equipments/{component_id}", response_model=BudgetAPUEquipment)
def update_equipment_in_item(budget_id: str, item_id: str, component_id: str, comp_in: BudgetAPUEquipmentUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    comp = db.query(DBEquipment).filter(DBEquipment.id == component_id, DBEquipment.budget_item_id == item_id).first()
    if not comp: raise HTTPException(status_code=404, detail="Equipment not found")
    for key, value in comp_in.dict(exclude_unset=True).items():
        setattr(comp, key, value)
    db.commit()
    db.refresh(comp)
    return comp

@router.put("/{budget_id}/items/{item_id}/labors/{component_id}", response_model=BudgetAPULabor)
def update_labor_in_item(budget_id: str, item_id: str, component_id: str, comp_in: BudgetAPULaborUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    comp = db.query(DBLabor).filter(DBLabor.id == component_id, DBLabor.budget_item_id == item_id).first()
    if not comp: raise HTTPException(status_code=404, detail="Labor not found")
    for key, value in comp_in.dict(exclude_unset=True).items():
        setattr(comp, key, value)
    db.commit()
    db.refresh(comp)
    return comp

@router.delete("/{budget_id}/items/{item_id}/materials/{component_id}")
def delete_material_from_item(budget_id: str, item_id: str, component_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    comp = db.query(DBMaterial).filter(DBMaterial.id == component_id, DBMaterial.budget_item_id == item_id).first()
    if not comp:
        raise HTTPException(status_code=404, detail="Material not found")
    db.delete(comp)
    db.commit()
    return {"ok": True}

@router.delete("/{budget_id}/items/{item_id}/equipments/{component_id}")
def delete_equipment_from_item(budget_id: str, item_id: str, component_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    comp = db.query(DBEquipment).filter(DBEquipment.id == component_id, DBEquipment.budget_item_id == item_id).first()
    if not comp:
        raise HTTPException(status_code=404, detail="Equipment not found")
    db.delete(comp)
    db.commit()
    return {"ok": True}

@router.delete("/{budget_id}/items/{item_id}/labors/{component_id}")
def delete_labor_from_item(budget_id: str, item_id: str, component_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    comp = db.query(DBLabor).filter(DBLabor.id == component_id, DBLabor.budget_item_id == item_id).first()
    if not comp:
        raise HTTPException(status_code=404, detail="Labor not found")
    db.delete(comp)
    db.commit()
    return {"ok": True}

@router.post("/{budget_id}/sync_prices")
def sync_budget_prices(budget_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    from app.db.models.cost360 import CostMaterial, CostEquipment, CostLabor
    
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
        
    for item in budget.items:
        # Sync Materials
        for mat in item.materials:
            cost_mat = db.query(CostMaterial).filter(CostMaterial.CodMat == mat.codigo).first()
            if cost_mat:
                mat.precio_unitario = cost_mat.CosMat if cost_mat.CosMat is not None else 0.0
                mat.descripcion = cost_mat.Descri if cost_mat.Descri is not None else mat.descripcion
                
        # Sync Equipment
        for eq in item.equipments:
            cost_eq = db.query(CostEquipment).filter(CostEquipment.CodEqu == eq.codigo).first()
            if cost_eq:
                eq.precio_unitario = cost_eq.CosDia if cost_eq.CosDia is not None else 0.0
                eq.descripcion = cost_eq.Descri if cost_eq.Descri is not None else eq.descripcion
                
        # Sync Labor
        for lab in item.labors:
            cost_lab = db.query(CostLabor).filter(CostLabor.CodMan == lab.codigo).first()
            if cost_lab:
                lab.jornal = cost_lab.Jornal if cost_lab.Jornal is not None else 0.0
                lab.bono = cost_lab.Bono if cost_lab.Bono is not None else 0.0
                lab.descripcion = cost_lab.Descri if cost_lab.Descri is not None else lab.descripcion
                
    db.commit()
    return {"status": "ok", "message": "Precios sincronizados con la Base Maestra"}

@router.post("/{budget_id}/export-excel")
async def export_budget_excel(budget_id: str, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    """Genera un archivo Excel con fórmulas para el presupuesto"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
        if not budget:
            raise HTTPException(status_code=404, detail="Budget not found")
        
        # Obtener items del presupuesto
        items = db.query(BudgetItem).filter(BudgetItem.budget_id == budget_id).all()
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto"
        
        # Estilos
        header_style = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_style = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        total_font = Font(bold=True, size=11)
        currency_format = '#,##0.00'
        
        # Header
        ws.merge_cells("B1:H1")
        ws["B1"] = budget.project_name or "PRESUPUESTO"
        ws["B1"].font = Font(bold=True, size=14)
        
        ws.merge_cells("B2:H2")
        ws["B2"] = f"Obra: {budget.project_name or ''}"
        
        if budget.client_name:
            ws.merge_cells("B3:H3")
            ws["B3"] = f"Contratante: {budget.client_name}"
        
        if budget.company_rif:
            ws.merge_cells("B4:H4")
            ws["B4"] = f"RIF: {budget.company_rif}"
        
        # Encabezados de tabla
        headers = ["Part. No", "Código COVENIN", "Descripción", "Unidad", "Cantidad", "Precio Unitario", "Total"]
        for i, header in enumerate(headers):
            cell = ws.cell(6, i + 2)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")
        
        # Filas de datos
        row_num = 7
        part_number = 1
        first_data_row = row_num
        
        for item in items:
            if not item.is_chapter:
                # Calcular precio unitario desde el APU del item
                pu = 0
                try:
                    # Intentar obtener APU para calcular precio unitario
                    from app.crud.crud_cost360 import get_apu_materials, get_apu_equipments, get_apu_labors
                    mat_rows = get_apu_materials(db, item.cod_par)
                    eq_rows = get_apu_equipments(db, item.cod_par)
                    mo_rows = get_apu_labors(db, item.cod_par)
                    
                    # Calcular precio unitario sumando materiales, equipos y mano de obra
                    total_mat = sum((mat.CosMat or 0) * (apu_mat.CanIns or 0) * (1 + (apu_mat.Desper or 0) / 100) 
                                  for apu_mat, mat in mat_rows)
                    total_eq = sum((eq.CosDia or 0) * (apu_eq.CanIns or 0) * (apu_eq.Deprec or 0) 
                                 for apu_eq, eq in eq_rows)
                    total_mo = sum((mo.Jornal or 0) * (apu_mo.CanIns or 0) + (mo.Bono or 0) * (apu_mo.CanIns or 0) 
                                 for apu_mo, mo in mo_rows)
                    
                    pu = round(total_mat + total_eq + total_mo, 2)
                except:
                    pu = 0
                
                total = pu * item.quantity
                
                ws.cell(row_num, 2, part_number)
                ws.cell(row_num, 3, item.cov_par or item.cod_par or '')
                ws.cell(row_num, 4, item.description)
                ws.cell(row_num, 5, item.unit)
                ws.cell(row_num, 6, item.quantity)
                ws.cell(row_num, 7, pu).number_format = currency_format
                ws.cell(row_num, 8, f"=G{row_num}*F{row_num}").number_format = currency_format
                
                part_number += 1
                row_num += 1
        
        last_data_row = row_num - 1
        
        # Totales
        ws.cell(row_num, 2, "Total (Sin I.V.A.):")
        ws.cell(row_num, 2).font = total_font
        ws.cell(row_num, 8, f"=SUM(H{first_data_row}:H{last_data_row})").number_format = currency_format
        ws.cell(row_num, 8).fill = total_style
        ws.cell(row_num, 8).font = total_font
        
        row_num += 1
        iva_percent = budget.iva_percent or 16
        ws.cell(row_num, 2, f"I.V.A. ({iva_percent}%):")
        ws.cell(row_num, 2).font = total_font
        ws.cell(row_num, 8, f"=H{row_num-1}*{iva_percent/100}").number_format = currency_format
        ws.cell(row_num, 8).font = total_font
        
        row_num += 1
        ws.cell(row_num, 2, "Total General:")
        ws.cell(row_num, 2).font = total_font
        ws.cell(row_num, 8, f"=H{row_num-2}+H{row_num-1}").number_format = currency_format
        ws.cell(row_num, 8).font = total_font
        
        # Ajustar anchos de columnas
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 60  # Descripción más ancha
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Guardar archivo temporal
        temp_dir = Path("temp")
        temp_dir.mkdir(exist_ok=True)
        import re
        filename = f"{re.sub(r'[^a-z0-9]', '_', budget.name.lower())}.xlsx"
        file_path = temp_dir / filename
        
        wb.save(file_path)
        
        return FileResponse(path=str(file_path), filename=filename)
    except Exception as e:
        import traceback
        print(f"Error exportando Excel: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error al exportar: {str(e)}")

@router.post("/{budget_id}/backup")
async def export_budget_backup(budget_id: str, request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    """Exporta un backup encriptado del presupuesto para descarga directa"""
    client_ip = request.client.host if request else None
    try:
        budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
        if not budget:
            log_backup_action(db, str(current_user.id), current_user.email, budget_id, "unknown", "export", "failed", "Budget not found", client_ip)
            raise HTTPException(status_code=404, detail="Budget not found")

        # Obtener items del presupuesto
        items = db.query(BudgetItem).filter(BudgetItem.budget_id == budget_id).all()

        # Serializar budget y items
        budget_dict = {
            "id": budget.id,
            "name": budget.name,
            "description": budget.description,
            "client_name": budget.client_name,
            "currency": budget.currency,
            "exchange_rate": budget.exchange_rate,
            "fcas_percent": budget.fcas_percent,
            "admin_percent": budget.admin_percent,
            "profit_percent": budget.profit_percent,
            "iva_percent": budget.iva_percent,
            "labor_bonus": budget.labor_bonus,
            "material_inflation": budget.material_inflation,
            "labor_inflation": budget.labor_inflation,
            "equipment_inflation": budget.equipment_inflation,
            "company_name": budget.company_name,
            "company_rif": budget.company_rif,
            "project_name": budget.project_name,
            "created_at": budget.created_at.isoformat() if budget.created_at else None,
            "updated_at": budget.updated_at.isoformat() if budget.updated_at else None
        }

        items_dict = []
        for item in items:
            item_data = {
                "id": item.id,
                "cod_par": item.cod_par,
                "cov_par": item.cov_par,
                "description": item.description,
                "unit": item.unit,
                "quantity": item.quantity,
                "performance": item.performance,
                "order": item.order,
                "is_chapter": item.is_chapter,
                "materials": [
                    {
                        "codigo": mat.codigo,
                        "descripcion": mat.descripcion,
                        "unidad": mat.unidad,
                        "precio_unitario": mat.precio_unitario,
                        "cantidad": mat.cantidad,
                        "desperdicio": mat.desperdicio
                    } for mat in item.materials
                ],
                "equipments": [
                    {
                        "codigo": eq.codigo,
                        "descripcion": eq.descripcion,
                        "unidad": eq.unidad,
                        "precio_unitario": eq.precio_unitario,
                        "cantidad": eq.cantidad,
                        "depreciacion": eq.depreciacion
                    } for eq in item.equipments
                ],
                "labors": [
                    {
                        "codigo": lab.codigo,
                        "descripcion": lab.descripcion,
                        "cantidad": lab.cantidad,
                        "jornal": lab.jornal,
                        "bono": lab.bono
                    } for lab in item.labors
                ]
            }
            items_dict.append(item_data)

        # Crear paquete de backup
        backup_package = encryption_service.create_backup_package(
            budget_dict, items_dict, current_user.email, str(current_user.id)
        )

        # Encriptar backup
        encrypted_backup = encryption_service.encrypt_backup(backup_package, current_user.email)

        # Registrar éxito en auditoría
        log_backup_action(db, str(current_user.id), current_user.email, budget_id, budget.name, "export", "success", None, client_ip)

        return {
            "status": "success",
            "message": f"Backup generado exitosamente para {budget.name}",
            "budget_name": budget.name,
            "backup_data": encrypted_backup.hex()  # Enviar como hex string
        }

    except Exception as e:
        import traceback
        print(f"Error exportando backup: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error al exportar backup: {str(e)}")

@router.post("/import-backup")
async def import_budget_backup(
    request: Request,
    backup_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_arko_admin)
):
    """Importa un backup encriptado de presupuesto"""
    client_ip = request.client.host if request else None
    budget_name_for_log = "unknown"
    try:
        # Validar que sea un archivo .cb
        if not backup_file.filename.endswith('.cb'):
            raise HTTPException(status_code=400, detail="Solo se permiten archivos .cb")

        # Leer el archivo encriptado
        encrypted_data = await backup_file.read()

        # Desencriptar el backup
        try:
            backup_data = encryption_service.decrypt_backup(encrypted_data, current_user.email)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Error desencriptando backup: {str(e)}")

        # Validar que el backup pertenezca al usuario
        if not encryption_service.validate_ownership(backup_data, current_user.email):
            raise HTTPException(status_code=403, detail="Este backup no pertenece a tu cuenta")

        # Validar versión del backup
        backup_version = backup_data.get('version')
        if backup_version != "1.0":
            raise HTTPException(status_code=400, detail="Versión de backup no compatible")

        # Extraer datos del presupuesto
        budget_data = backup_data.get('budget')
        items_data = backup_data.get('items', [])
        budget_name_for_log = budget_data.get('name', 'unknown')

        if not budget_data:
            raise HTTPException(status_code=400, detail="Backup corrupto: no contiene datos de presupuesto")

        # Crear nuevo presupuesto (no mantener el ID original)
        new_budget = Budget(
            user_id=str(current_user.id),
            name=budget_data.get('name', 'Importado desde backup'),
            description=budget_data.get('description'),
            client_name=budget_data.get('client_name'),
            currency=budget_data.get('currency', 'USD'),
            exchange_rate=budget_data.get('exchange_rate', 1.0),
            fcas_percent=budget_data.get('fcas_percent', 417.0),
            admin_percent=budget_data.get('admin_percent', 15.0),
            profit_percent=budget_data.get('profit_percent', 10.0),
            iva_percent=budget_data.get('iva_percent', 16.0),
            labor_bonus=budget_data.get('labor_bonus', 0.0),
            material_inflation=budget_data.get('material_inflation', 0.0),
            labor_inflation=budget_data.get('labor_inflation', 0.0),
            equipment_inflation=budget_data.get('equipment_inflation', 0.0),
            company_name=budget_data.get('company_name'),
            company_rif=budget_data.get('company_rif'),
            project_name=budget_data.get('project_name')
        )

        db.add(new_budget)
        db.commit()
        db.refresh(new_budget)

        # Importar items
        max_order = 0
        for item_data in items_data:
            target_order = item_data.get('order', 0)
            if target_order <= 0:
                max_order += 1
                target_order = max_order
            else:
                max_order = max(max_order, target_order)

            # Crear el item
            new_item = BudgetItem(
                budget_id=new_budget.id,
                cod_par=item_data.get('cod_par', ''),
                cov_par=item_data.get('cov_par'),
                description=item_data.get('description', ''),
                unit=item_data.get('unit', ''),
                quantity=item_data.get('quantity', 0.0),
                performance=item_data.get('performance', 1.0),
                order=target_order,
                is_chapter=item_data.get('is_chapter', False)
            )

            db.add(new_item)
            db.commit()
            db.refresh(new_item)

            # Si no es capítulo, importar materiales, equipos y mano de obra
            if not item_data.get('is_chapter', False):
                # Importar materiales
                for mat_data in item_data.get('materials', []):
                    db_mat = DBMaterial(
                        budget_item_id=new_item.id,
                        codigo=mat_data.get('codigo', ''),
                        descripcion=mat_data.get('descripcion', ''),
                        unidad=mat_data.get('unidad', ''),
                        precio_unitario=mat_data.get('precio_unitario', 0.0),
                        cantidad=mat_data.get('cantidad', 0.0),
                        desperdicio=mat_data.get('desperdicio', 0.0)
                    )
                    db.add(db_mat)

                # Importar equipos
                for eq_data in item_data.get('equipments', []):
                    db_eq = DBEquipment(
                        budget_item_id=new_item.id,
                        codigo=eq_data.get('codigo', ''),
                        descripcion=eq_data.get('descripcion', ''),
                        unidad=eq_data.get('unidad', ''),
                        precio_unitario=eq_data.get('precio_unitario', 0.0),
                        cantidad=eq_data.get('cantidad', 0.0),
                        depreciacion=eq_data.get('depreciacion', 1.0)
                    )
                    db.add(db_eq)

                # Importar mano de obra
                for lab_data in item_data.get('labors', []):
                    db_lab = DBLabor(
                        budget_item_id=new_item.id,
                        codigo=lab_data.get('codigo', ''),
                        descripcion=lab_data.get('descripcion', ''),
                        cantidad=lab_data.get('cantidad', 0.0),
                        jornal=lab_data.get('jornal', 0.0),
                        bono=lab_data.get('bono', 0.0)
                    )
                    db.add(db_lab)

        db.commit()
        db.refresh(new_budget)

        # Registrar éxito en auditoría
        log_backup_action(db, str(current_user.id), current_user.email, new_budget.id, new_budget.name, "import", "success", None, client_ip)

        return {
            "status": "success",
            "message": "Backup importado exitosamente",
            "budget_id": new_budget.id,
            "budget_name": new_budget.name
        }

    except HTTPException as e:
        # Registrar error en auditoría
        log_backup_action(db, str(current_user.id), current_user.email, "unknown", budget_name_for_log, "import", "failed", str(e.detail), client_ip)
        raise
    except Exception as e:
        import traceback
        print(f"Error importando backup: {str(e)}")
        print(traceback.format_exc())
        # Registrar error en auditoría
        log_backup_action(db, str(current_user.id), current_user.email, "unknown", budget_name_for_log, "import", "failed", str(e), client_ip)
        raise HTTPException(status_code=500, detail=f"Error al importar backup: {str(e)}")

@router.post("/{budget_id}/upload-logo")
async def upload_budget_logo(budget_id: str, logo: UploadFile = File(...), db: Session = Depends(get_db)):
    """Sube el logo de la empresa para un presupuesto específico"""
    budget = db.query(Budget).filter(Budget.id == budget_id, Budget.user_id == str(current_user.id)).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    
    # Crear directorio para logos si no existe
    upload_dir = Path("public/company_logos")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Generar nombre único para el archivo
    file_extension = logo.filename.split('.')[-1].lower()
    if file_extension not in ['png', 'jpg', 'jpeg']:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PNG, JPG o JPEG")
    
    unique_filename = f"{budget.id}_{uuid.uuid4().hex[:8]}.{file_extension}"
    file_path = upload_dir / unique_filename
    
    # Guardar el archivo
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(logo.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error guardando el archivo: {str(e)}")
    
    # Actualizar el presupuesto con la URL del logo
    logo_url = f"/company_logos/{unique_filename}"
    budget.company_logo = logo_url
    db.commit()
    
    return {"logo_url": logo_url}