from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import Optional
from pathlib import Path
import openpyxl

from app.db.base import get_db
from app.api.v1.endpoints.arko import get_current_arko_admin
from app.middleware.plan_limits import check_ai_access
from app.schemas.cost360 import (
    CostItemListResponse, APUResponse, APUComponent,
    CostMaterialUpdate, CostEquipmentUpdate, CostLaborUpdate,
    AiApuGenerateRequest, SmartSelectRequest,
    CustomCostItemCreate, CustomCostItemResponse,
    Cost360DatabaseCreate, Cost360DatabaseUpdate, Cost360DatabaseListResponse,
    MasterItemUpdate, MasterAPUUpdate, CustomApuExportRequest
)
import re
from sqlalchemy import text

def set_schema_for_db(db: Session, database_id: str):
    if database_id and database_id not in ["master", "personalizada"] and re.match(r'^[a-zA-Z0-9_\-]+$', database_id):
        try:
            db.execute(text(f'SET LOCAL search_path TO "{database_id}", public'))
        except:
            pass

# Import Services and CRUD
from app.crud.crud_cost360 import (
    get_items_paginated, get_item_by_code,
    get_apu_materials, get_apu_equipments, get_apu_labors,
    search_materials_paginated, search_equipments_paginated, search_labors_paginated,
    get_categories_tree_data,
    update_material, delete_material,
    update_equipment, delete_equipment,
    update_labor, delete_labor,
    save_custom_apu,
    get_all_databases, get_database_by_id, create_database, update_database, delete_database,
    update_master_item, delete_master_item, update_master_apu_details
)
from app.services.preprocessing_service import preprocess_apu_data, fast_preprocess_debug
from app.services.ai_apu_service import generate_apu_with_ai, generate_apu_with_ai_from_base
from app.api.v1.endpoints.export_utils import generate_excel_workbook

router = APIRouter()

@router.get("/items", response_model=CostItemListResponse)
def get_items(skip: int = 0, limit: int = 50, search: Optional[str] = None, chapter: Optional[str] = None, categoria: Optional[str] = None, tipo_actividad: Optional[str] = None, search_desc: bool = True, search_insumos: bool = False, covenin: Optional[str] = None, database_id: str = "master", only_coded: bool = False, hidden_categories: Optional[str] = None, db: Session = Depends(get_db)):
    if database_id.startswith("budget_"):
        budget_id = database_id.replace("budget_", "")
        from app.db.models.budget import BudgetItem
        from sqlalchemy import or_, func
        
        query = db.query(BudgetItem).filter(BudgetItem.budget_id == budget_id, BudgetItem.is_chapter == False)
        
        if search:
            search_term = f"%{search.lower()}%"
            query = query.filter(
                or_(
                    func.lower(BudgetItem.cod_par).like(search_term),
                    func.lower(BudgetItem.description).like(search_term)
                )
            )
            
        total = query.count()
        budget_items = query.order_by(BudgetItem.order).offset(skip).limit(limit).all()
        
        items = []
        for bi in budget_items:
            items.append({
                "CodPar": bi.cod_par,
                "Descri": bi.description,
                "CovPar": bi.cov_par,
                "UniPar": bi.unit,
                "PreUni": 0.0,
                "RenPar": bi.performance
            })
            
        return {"total": total, "items": items}

    set_schema_for_db(db, database_id)
    total, items = get_items_paginated(db, skip, limit, search, chapter, categoria, tipo_actividad, search_desc, search_insumos, covenin, database_id, only_coded, hidden_categories)
    return {"total": total, "items": items}


def _get_db_factors(db: Session, database_id: str) -> dict:
    """Obtener los factores de inflación de una base de datos por su ID."""
    if not database_id or database_id == 'master':
        return {"mat": 1.0, "lab": 1.0, "eq": 1.0}
    db_config = get_database_by_id(db, database_id)
    if not db_config:
        return {"mat": 1.0, "lab": 1.0, "eq": 1.0}
    return {
        "mat": 1 + (db_config.material_inflation or 0.0) / 100.0,
        "lab": 1 + (db_config.labor_inflation or 0.0) / 100.0,
        "eq": 1 + (db_config.equipment_inflation or 0.0) / 100.0,
    }

@router.get("/items/{item_code}/apu", response_model=APUResponse)
def get_apu(item_code: str, database_id: str = "master", db: Session = Depends(get_db)):
    if database_id.startswith("budget_"):
        budget_id = database_id.replace("budget_", "")
        from app.db.models.budget import BudgetItem
        bi = db.query(BudgetItem).filter(BudgetItem.budget_id == budget_id, BudgetItem.cod_par == item_code).first()
        if not bi:
            raise HTTPException(status_code=404, detail="Partida de presupuesto no encontrada")
            
        partida = {
            "CodPar": bi.cod_par,
            "Descri": bi.description,
            "CovPar": bi.cov_par,
            "UniPar": bi.unit,
            "PreUni": 0.0,
            "RenPar": bi.performance
        }
        
        materials = [
            APUComponent(codigo=m.codigo, descripcion=m.descripcion, unidad=m.unidad, cantidad=m.cantidad, precio_unitario=m.precio_unitario, subtotal=m.cantidad*m.precio_unitario*(1+m.desperdicio/100), desperdicio=m.desperdicio) for m in bi.materials
        ]
        equipments = [
            APUComponent(codigo=e.codigo, descripcion=e.descripcion, unidad=e.unidad, cantidad=e.cantidad, precio_unitario=e.precio_unitario, subtotal=e.cantidad*e.precio_unitario*(e.depreciacion), depreciacion=e.depreciacion) for e in bi.equipments
        ]
        labors = [
            APUComponent(codigo=l.codigo, descripcion=l.descripcion, unidad="Día", cantidad=l.cantidad, precio_unitario=l.jornal+l.bono, subtotal=l.cantidad*(l.jornal+l.bono), jornal=l.jornal, bono=l.bono, tot_jornal=l.cantidad*l.jornal, tot_bono=l.cantidad*l.bono) for l in bi.labors
        ]
        
        total_directo = sum(c.subtotal for c in materials) + sum(c.subtotal for c in equipments) + sum(c.subtotal for c in labors)
        return {"partida": partida, "materiales": materials, "equipos": equipments, "mano_obra": labors, "total_directo": total_directo}

    set_schema_for_db(db, database_id)
    if item_code.startswith("CUST-"):
        from app.db.models.cost360 import CustomCostItem
        import json
        custom_items = db.query(CustomCostItem).all()
        for ci in custom_items:
            try:
                data = json.loads(ci.apu_data)
                cod = data.get("cod_par") or ("CUST-" + ci.id[:4].upper())
                if cod == item_code:
                    partida = {
                        "CodPar": cod,
                        "Descri": ci.description,
                        "UniPar": ci.unit,
                        "RenPar": ci.performance
                    }
                    materials = [
                        APUComponent(codigo=m.get('id',''), descripcion=m.get('descripcion',''), unidad=m.get('unidad',''), cantidad=m.get('cantidad',0), precio_unitario=m.get('precio_unitario',0), subtotal=m.get('cantidad',0)*m.get('precio_unitario',0)*(1+m.get('desperdicio',0)/100), desperdicio=m.get('desperdicio',0)) for m in data.get('materials', [])
                    ]
                    equipments = [
                        APUComponent(codigo=e.get('id',''), descripcion=e.get('descripcion',''), unidad=e.get('unidad',''), cantidad=e.get('cantidad',0), precio_unitario=e.get('precio_unitario',0), subtotal=e.get('cantidad',0)*e.get('precio_unitario',0)*(e.get('depreciacion',1.0)), depreciacion=e.get('depreciacion',1.0)) for e in data.get('equipments', [])
                    ]
                    labors = [
                        APUComponent(codigo=l.get('id',''), descripcion=l.get('descripcion',''), unidad=l.get('unidad',''), cantidad=l.get('cantidad',0), precio_unitario=l.get('jornal',0), subtotal=l.get('cantidad',0)*l.get('jornal',0), jornal=l.get('jornal',0), bono=l.get('bono',0)) for l in data.get('labor', data.get('labors', []))
                    ]
                    total_directo = sum(c.subtotal for c in materials) + sum(c.subtotal for c in equipments) + sum(c.subtotal for c in labors)
                    return {"partida": partida, "materiales": materials, "equipos": equipments, "mano_obra": labors, "total_directo": total_directo}
            except:
                continue
        raise HTTPException(status_code=404, detail="Partida personalizada no encontrada")

    item = get_item_by_code(db, item_code)
    if not item:
        raise HTTPException(status_code=404, detail="Partida no encontrada")

    factors = _get_db_factors(db, database_id)

    mat_results = get_apu_materials(db, item_code)
    materiales = []
    for rel, mat in mat_results:
        desperdicio = rel.Desper if hasattr(rel, 'Desper') and rel.Desper else 0.0
        precio = (mat.CosMat or 0.0) * factors["mat"]
        subtotal = rel.CanIns * precio * (1 + (desperdicio / 100.0))
        materiales.append(APUComponent(
            codigo=mat.CodMat, descripcion=mat.Descri, unidad=mat.UniMat, cantidad=rel.CanIns,
            precio_unitario=round(precio, 4), subtotal=round(subtotal, 2), desperdicio=desperdicio
        ))

    eq_results = get_apu_equipments(db, item_code)
    equipos = []
    for rel, eq in eq_results:
        depreciacion = rel.Deprec if hasattr(rel, 'Deprec') and rel.Deprec else 1.0
        precio_diario_depreciado = (eq.CosDia or 0.0) * factors["eq"]
        precio_adquisicion = precio_diario_depreciado / depreciacion if depreciacion > 0 else precio_diario_depreciado
        subtotal = rel.CanIns * precio_diario_depreciado
        equipos.append(APUComponent(
            codigo=eq.CodEqu, descripcion=eq.Descri, unidad="Día", cantidad=rel.CanIns,
            precio_unitario=round(precio_adquisicion, 4), subtotal=round(subtotal, 2), depreciacion=depreciacion
        ))

    mo_results = get_apu_labors(db, item_code)
    mano_obra = []
    for rel, mo in mo_results:
        jornal = (mo.Jornal or 0.0) * factors["lab"]
        bono = (mo.Bono or 0.0) * factors["lab"]
        tot_jornal = rel.CanIns * jornal
        tot_bono = rel.CanIns * bono
        precio = jornal + bono
        subtotal = tot_jornal + tot_bono
        mano_obra.append(APUComponent(
            codigo=mo.CodMan, descripcion=mo.Descri, unidad="Día", cantidad=rel.CanIns,
            precio_unitario=round(precio, 2), subtotal=round(subtotal, 2),
            jornal=round(jornal, 4), bono=round(bono, 4),
            tot_jornal=round(tot_jornal, 2), tot_bono=round(tot_bono, 2)
        ))

    total_directo = sum(c.subtotal for c in materiales) + sum(c.subtotal for c in equipos) + sum(c.subtotal for c in mano_obra)

    return APUResponse(
        partida=item, materiales=materiales, equipos=equipos, mano_obra=mano_obra, total_directo=round(total_directo, 2)
    )

@router.put("/items/{item_code}")
def update_master_item_route(item_code: str, payload: MasterItemUpdate, db: Session = Depends(get_db)):
    updated_item = update_master_item(db, item_code, payload.Descri, payload.UniPar, payload.RenPar)
    if not updated_item:
        raise HTTPException(status_code=404, detail="Partida no encontrada")
    return updated_item

@router.put("/items/{item_code}/apu")
def update_master_apu_route(item_code: str, payload: MasterAPUUpdate, database_id: str = "master", db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    set_schema_for_db(db, database_id)
    updated_item = update_master_apu_details(
        db=db,
        item_code=item_code,
        description=payload.description,
        unit=payload.unit,
        performance=payload.performance,
        materials=[m.model_dump() for m in payload.materials] if payload.materials is not None else None,
        equipments=[e.model_dump() for e in payload.equipments] if payload.equipments is not None else None,
        labors=[l.model_dump() for l in payload.labors] if payload.labors is not None else None
    )
    if not updated_item:
        raise HTTPException(status_code=404, detail="Partida no encontrada")
    return {
        "status": "ok",
        "message": "APU actualizado correctamente",
        "item": {
            "CodPar": updated_item.CodPar,
            "CovPar": updated_item.CovPar,
            "Descri": updated_item.Descri,
            "UniPar": updated_item.UniPar,
            "RenPar": updated_item.RenPar,
            "PreUni": updated_item.PreUni
        }
    }

@router.delete("/items/{item_code}")
def delete_master_item_route(item_code: str, db: Session = Depends(get_db)):
    if not delete_master_item(db, item_code):
        raise HTTPException(status_code=404, detail="Partida no encontrada")
    return {"status": "ok"}

@router.get("/materials")
def search_materials_route(skip: int = 0, limit: int = 50, search: str = "", database_id: str = "master", db: Session = Depends(get_db)):
    set_schema_for_db(db, database_id)
    total, items = search_materials_paginated(db, skip, limit, search)
    # Aplicar factor de inflación de materiales si la base no es maestra
    if database_id and database_id != "master":
        db_config = get_database_by_id(db, database_id)
        if db_config and db_config.material_inflation:
            factor = 1 + (db_config.material_inflation / 100.0)
            for item in items:
                item.CosMat = round((item.CosMat or 0.0) * factor, 4)
    return {"total": total, "items": items}

@router.get("/equipments")
def search_equipments_route(skip: int = 0, limit: int = 50, search: str = "", database_id: str = "master", db: Session = Depends(get_db)):
    set_schema_for_db(db, database_id)
    total, items = search_equipments_paginated(db, skip, limit, search)
    # Aplicar factor de inflación de equipos si la base no es maestra
    if database_id and database_id != "master":
        db_config = get_database_by_id(db, database_id)
        if db_config and db_config.equipment_inflation:
            factor = 1 + (db_config.equipment_inflation / 100.0)
            for item in items:
                item.CosDia = round((item.CosDia or 0.0) * factor, 4)
    return {"total": total, "items": items}

@router.get("/labors")
def search_labors_route(skip: int = 0, limit: int = 50, search: str = "", database_id: str = "master", db: Session = Depends(get_db)):
    set_schema_for_db(db, database_id)
    total, items = search_labors_paginated(db, skip, limit, search)
    # Aplicar factor de inflación de mano de obra si la base no es maestra
    if database_id and database_id != "master":
        db_config = get_database_by_id(db, database_id)
        if db_config and db_config.labor_inflation:
            factor = 1 + (db_config.labor_inflation / 100.0)
            for item in items:
                item.Jornal = round((item.Jornal or 0.0) * factor, 4)
                item.Bono = round((item.Bono or 0.0) * factor, 4)
    return {"total": total, "items": items}

@router.get("/categories_tree")
def get_categories_tree_route(db: Session = Depends(get_db)):
    return get_categories_tree_data(db)

@router.patch("/materials/{codigo}")
def update_material_route(codigo: str, payload: CostMaterialUpdate, db: Session = Depends(get_db)):
    mat = update_material(db, codigo, payload)
    if not mat: raise HTTPException(status_code=404, detail="Material no encontrado")
    return mat

@router.post("/materials/bulk-update")
def bulk_update_materials(payload: dict, db: Session = Depends(get_db)):
    """
    Actualización masiva de precios - NO consume tokens de IA
    Actualización directa en base de datos para máxima eficiencia
    """
    try:
        updates = payload.get('updates', [])
        if not updates:
            return {"updated": 0, "errors": [], "total": 0}
        
        # Optimización: usar UPDATE directo en lugar de queries individuales
        updated_count = 0
        errors = []
        
        # Preparar datos para actualización en lote
        codigos_precio = {}
        for update in updates:
            codigo = update.get('codigo')
            precio = update.get('precio')
            if codigo and precio is not None:
                codigos_precio[codigo] = precio
        
        # Actualización en lote para mejor rendimiento
        for codigo, precio in codigos_precio.items():
            try:
                result = db.execute(
                    text('UPDATE cost360_materials SET "CosMat" = :precio WHERE "CodMat" = :codigo'),
                    {"precio": precio, "codigo": codigo}
                )
                if result.rowcount > 0:
                    updated_count += result.rowcount
                else:
                    errors.append(f"Material {codigo} no encontrado")
            except Exception as e:
                errors.append(f"Error actualizando {codigo}: {str(e)}")
        
        db.commit()
        
        return {
            "updated": updated_count,
            "errors": errors,
            "total": len(updates)
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error en actualización masiva: {str(e)}")

@router.post("/materials/bulk-update-descriptions")
async def bulk_update_descriptions(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Actualización masiva de descripciones desde archivo Excel
    Formato esperado: columnas "Código" y "Descripción"
    """
    try:
        # Leer el archivo Excel con openpyxl
        import io
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Encontrar índices de columnas requeridas
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        codigo_col_idx = None
        descripcion_col_idx = None

        for idx, header in enumerate(header_row):
            if 'Código' in str(header) or 'Codigo' in str(header):
                codigo_col_idx = idx
            elif 'Descripción' in str(header) or 'Descripcion' in str(header) or 'Descri' in str(header):
                descripcion_col_idx = idx

        if codigo_col_idx is None or descripcion_col_idx is None:
            raise HTTPException(
                status_code=400,
                detail=f"Columnas requeridas no encontradas. Se necesita 'Código' y 'Descripción'. Encabezados encontrados: {header_row}"
            )

        # Procesar actualizaciones
        updated_count = 0
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(codigo_col_idx, descripcion_col_idx):
                continue

            codigo = str(row[codigo_col_idx]).strip() if row[codigo_col_idx] else ""
            descripcion = str(row[descripcion_col_idx]).strip() if row[descripcion_col_idx] else ""

            if not codigo or not descripcion:
                continue

            try:
                result = db.execute(
                    text('UPDATE cost360_materials SET "Descri" = :descripcion WHERE "CodMat" = :codigo'),
                    {"descripcion": descripcion, "codigo": codigo}
                )
                if result.rowcount > 0:
                    updated_count += result.rowcount
                else:
                    errors.append(f"Material {codigo} no encontrado")
            except Exception as e:
                errors.append(f"Error actualizando {codigo}: {str(e)}")

        db.commit()

        return {
            "updated": updated_count,
            "errors": errors,
            "total": ws.max_row - 1  # Excluyendo header
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error en actualización masiva de descripciones: {str(e)}")

@router.delete("/materials/{codigo}")
def delete_material_route(codigo: str, db: Session = Depends(get_db)):
    if not delete_material(db, codigo): raise HTTPException(status_code=404, detail="Material no encontrado")
    return {"status": "ok"}

@router.patch("/equipments/{codigo}")
def update_equipment_route(codigo: str, payload: CostEquipmentUpdate, db: Session = Depends(get_db)):
    eq = update_equipment(db, codigo, payload)
    if not eq: raise HTTPException(status_code=404, detail="Equipo no encontrado")
    return eq

@router.delete("/equipments/{codigo}")
def delete_equipment_route(codigo: str, db: Session = Depends(get_db)):
    if not delete_equipment(db, codigo): raise HTTPException(status_code=404, detail="Equipo no encontrado")
    return {"status": "ok"}

@router.patch("/labors/{codigo}")
def update_labor_route(codigo: str, payload: CostLaborUpdate, db: Session = Depends(get_db)):
    labor = update_labor(db, codigo, payload)
    if not labor: raise HTTPException(status_code=404, detail="Mano de obra no encontrada")
    return labor

@router.delete("/labors/{codigo}")
def delete_labor_route(codigo: str, db: Session = Depends(get_db)):
    if not delete_labor(db, codigo): raise HTTPException(status_code=404, detail="Mano de obra no encontrada")
    return {"status": "ok"}

@router.post("/generate-ai-apu")
def generate_ai_apu_route(payload: AiApuGenerateRequest, db: Session = Depends(get_db), current_user = Depends(get_current_arko_admin)):
    # Verificar acceso a IA
    check_ai_access(current_user)

    # 0. Si es solo preproceso DEBUG, devolver resultado rapido
    if payload.only_preprocess:
        from app.services.ai_search import ai_engine
        debug_data = fast_preprocess_debug(
            db, payload.description, payload.covenin_prefix, payload.covenin_context
        )
        # Inyectar estado real del motor IA para diagnóstico
        debug_data["motor_ia_estado"] = {
            "is_loaded": ai_engine.is_loaded,
            "total_ids_mapeados": len(ai_engine.ids_mapping),
            "embeddings_forma": str(ai_engine.embeddings.shape) if ai_engine.embeddings is not None else "No cargado",
        }
        return {
            "status": "clarification_needed",
            "clarification_message": f"MODO DEBUG: {len(debug_data.get('todas_las_partidas_covenin', []))} candidatas encontradas tras expansión dinámica",
            "options": [],
            "questions": [],
            "debug_preprocesamiento": debug_data
        }

    # 0.5. MODO SMART SELECTOR: el usuario ya eligió una partida base -> generar con base real
    if payload.base_partida_code:
        from app.services.smart_selector_service import fetch_base_apu_for_prompt, _get_dynamic_candidates
        import logging
        logger = logging.getLogger(__name__)
        
        base_apu = fetch_base_apu_for_prompt(db, payload.base_partida_code)
        
        complementary_apus = []
        try:
            candidates, _ = _get_dynamic_candidates(db, payload.description, payload.covenin_prefix or "", limit=4)
            comp_items = [c["item"] for c in candidates if c["item"].CodPar != payload.base_partida_code][:2]
            for item in comp_items:
                comp_apu = fetch_base_apu_for_prompt(db, item.CodPar)
                if comp_apu:
                    complementary_apus.append(comp_apu)
        except Exception as e:
            logger.warning(f"Error fetching complementary APUs: {e}")

        history_dicts = [msg.model_dump() for msg in payload.history] if payload.history else []
        result = generate_apu_with_ai_from_base(
            base_apu=base_apu,
            complementary_apus=complementary_apus,
            user_description=payload.description,
            covenin_prefix=payload.covenin_prefix or "",
            covenin_context=payload.covenin_context or "",
            smart_answers=payload.smart_answers or {},
            history=history_dicts,
        )
        if (result.get("status") in ("success", "completed")) and result.get("partida"):
            from app.db.arko_base import ArkoSessionLocal
            with ArkoSessionLocal() as adb:
                db_user = adb.query(current_user.__class__).filter_by(id=current_user.id).first()
                if db_user:
                    db_user.ai_apus_generated = getattr(db_user, 'ai_apus_generated', 0) + 1
                    adb.commit()
        return result

    # 1. Preprocesamiento (BD + Estadísticas) + IA semantica
    payload_llm = preprocess_apu_data(db, payload.description, payload.covenin_prefix, payload.covenin_context)
    
    # 0.8. Si el usuario aceptó la partida de Match Exacto ("Sí, es esa")
    if payload.accept_exact_match_code:
        item = get_item_by_code(db, payload.accept_exact_match_code)
        if item:
            mat_results = get_apu_materials(db, item.CodPar)
            eq_results = get_apu_equipments(db, item.CodPar)
            mo_results = get_apu_labors(db, item.CodPar)

            materials = [
                {
                    "id": f"m-{mat.CodMat}",
                    "codigo": mat.CodMat,
                    "descripcion": mat.Descri,
                    "unidad": mat.UniMat,
                    "cantidad": rel.CanIns,
                    "desperdicio": getattr(rel, 'Desper', 0.0) or 0.0,
                    "precio_unitario": mat.CosMat or 0.0,
                    "origen": "historico",
                    "nota_calculo": "Extraído directamente de la base de datos certificada."
                } for rel, mat in mat_results
            ]

            equipments = [
                {
                    "id": f"e-{eq.CodEqu}",
                    "codigo": eq.CodEqu,
                    "descripcion": eq.Descri,
                    "unidad": "día",
                    "cantidad": rel.CanIns,
                    "depreciacion": getattr(rel, 'Deprec', 1.0) or 1.0,
                    "precio_unitario": eq.CosDia or 0.0,
                    "origen": "historico",
                    "nota_calculo": "Extraído directamente de la base de datos certificada."
                } for rel, eq in eq_results
            ]

            labors = [
                {
                    "id": f"l-{mo.CodMan}",
                    "codigo": mo.CodMan,
                    "descripcion": mo.Descri,
                    "unidad": "día",
                    "cantidad": rel.CanIns,
                    "jornal": mo.Jornal or 0.0,
                    "bono": mo.Bono or 0.0,
                    "precio_unitario": (mo.Jornal or 0.0) + (mo.Bono or 0.0),
                    "origen": "historico",
                    "nota_calculo": "Extraído directamente de la base de datos certificada."
                } for rel, mo in mo_results
            ]

            return {
                "status": "completed",
                "partida": {
                    "cod_par": item.CodPar,
                    "cov_par": item.CovPar or item.CodPar,
                    "description": item.Descri,
                    "unit": item.UniPar,
                    "quantity": 1.0,
                    "performance": getattr(item, 'RenPar', 1.0) or 1.0
                },
                "materials": materials,
                "equipments": equipments,
                "labors": labors,
                "advertencias": [
                    f"Partida certificada [{item.CodPar}] importada directamente desde la base de datos maestra a solicitud del usuario."
                ]
            }

    # 1. Preprocesamiento (BD + Estadísticas) + IA semantica
    payload_llm = preprocess_apu_data(db, payload.description, payload.covenin_prefix, payload.covenin_context)
    
    # 1.5. Pregunta interactiva si hay Match Exacto y el usuario aún no ha omitido
    if payload_llm.get("modo") == "partida_exacta_encontrada" and not payload.bypass_exact_match:
        cod_par = payload_llm.get("partida_exacta_codigo")
        item = get_item_by_code(db, cod_par)
        if item:
            return {
                "status": "exact_match_candidate",
                "matched_item": {
                    "cod_par": item.CodPar,
                    "cov_par": item.CovPar or item.CodPar,
                    "description": item.Descri,
                    "unit": item.UniPar,
                    "pre_uni": item.PreUni or 0.0,
                    "performance": getattr(item, 'RenPar', 1.0) or 1.0
                },
                "message": "Existe una partida que coincide casi al 100% con tu descripción:"
            }

    # 2. Generación con IA (LLM Router)
    history_dicts = [msg.model_dump() for msg in payload.history] if payload.history else []
    result = generate_apu_with_ai(payload_llm, history_dicts)
    
    if (result.get("status") in ("success", "completed")) and result.get("partida"):
        from app.db.arko_base import ArkoSessionLocal
        with ArkoSessionLocal() as adb:
            db_user = adb.query(current_user.__class__).filter_by(id=current_user.id).first()
            if db_user:
                db_user.ai_apus_generated = getattr(db_user, 'ai_apus_generated', 0) + 1
                adb.commit()

    return result


@router.post("/smart-select")
def smart_select_route(payload: SmartSelectRequest, db: Session = Depends(get_db)):
    """Selección guiada de partida base mediante preguntas discriminantes. Sin LLM."""
    from app.services.smart_selector_service import get_smart_selector_data
    return get_smart_selector_data(
        db=db,
        description=payload.description,
        covenin_prefix=payload.covenin_prefix,
        covenin_context=payload.covenin_context,
        answers=payload.answers or {},
    )


@router.post("/custom-apus", response_model=CustomCostItemResponse)
def save_custom_apu_route(payload: CustomCostItemCreate, db: Session = Depends(get_db)):
    new_item = save_custom_apu(db, payload.description, payload.unit, payload.performance, payload.apu_data)
    return new_item




@router.post("/apu/{item_id}/export-excel")
async def export_apu_excel(item_id: str, db: Session = Depends(get_db)):
    """Genera un archivo Excel con fórmulas nativas usando el formato del script de referencia apu_formulas.py"""
    try:
        # Obtener la partida principal
        try:
            item = get_item_by_code(db, item_id.split('-')[0])
        except Exception:
            raise HTTPException(status_code=404, detail="Item not found")

        # Obtener APU
        mat_rows = get_apu_materials(db, item_id)
        eq_rows = get_apu_equipments(db, item_id)
        mo_rows = get_apu_labors(db, item_id)

        # Convertir a dicts
        item_dict = {
            "CodPar": item.CodPar,
            "CovPar": item.CovPar,
            "Descri": item.Descri,
            "UniPar": item.UniPar,
            "RenPar": item.RenPar
        }
        
        mats = [{"Descri": mat.Descri if mat else '', "UniMat": mat.UniMat if mat else '', "CanIns": apu.CanIns, "Desper": apu.Desper, "CosMat": mat.CosMat if mat else 0} for apu, mat in mat_rows]
        eqs = [{"Descri": eq.Descri if eq else '', "CanIns": apu.CanIns, "Deprec": apu.Deprec, "CosDia": eq.CosDia if eq else 0} for apu, eq in eq_rows]
        mos = [{"Descri": mo.Descri if mo else '', "CanIns": apu.CanIns, "Jornal": mo.Jornal if mo else 0, "Bono": mo.Bono if mo else 0} for apu, mo in mo_rows]

        file_path, filename = generate_excel_workbook(item_dict, mats, eqs, mos)
        return FileResponse(path=str(file_path), filename=filename)

    except Exception as e:
        import traceback
        print(f"Error exportando APU Excel: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error al exportar APU: {str(e)}")
@router.post("/apu/export-excel-custom")
async def export_apu_excel_custom(payload: CustomApuExportRequest):
    """Genera un archivo Excel desde la memoria enviada por el frontend (APU dinámico o en edición)"""
    try:
        # Extraer data
        item_data = payload.item
        mats = payload.materials
        eqs = payload.equipments
        mos = payload.labors
        
        # Mapear a las llaves que espera export_utils
        item_dict = {
            "CodPar": item_data.get("CodPar") or item_data.get("cod_par", "Custom"),
            "Descri": item_data.get("Descri") or item_data.get("description", "Custom APU"),
            "UniPar": item_data.get("UniPar") or item_data.get("unit", "UND"),
            "RenPar": item_data.get("RenPar") or item_data.get("performance", 1.0)
        }
        
        settings = payload.settings or {}
        
        file_path, filename = generate_excel_workbook(item_dict, mats, eqs, mos, settings)
        
        return FileResponse(
            path=str(file_path), 
            filename=filename, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        import traceback
        print(f"Error exportando APU custom a Excel: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error al exportar APU: {str(e)}")




@router.post("/rag/update-brain")
def update_rag_brain(background_tasks: BackgroundTasks):
    """
    Ejecuta la actualización del Cerebro RAG (generación de embeddings y CSV) en segundo plano.
    Este proceso lee las partidas de PostgreSQL, calcula los embeddings con MiniLM y 
    los guarda en la carpeta /app para que el AISearchEngine los cargue en el proximo restart.
    """
    import subprocess
    
    def run_generation():
        try:
            # Ejecutamos el script que ya existe en el contenedor
            subprocess.run(
                ["python3", "/app/generate_embeddings.py"], 
                capture_output=True, 
                text=True, 
                check=True
            )
            print("Generación de Cerebro RAG finalizada con éxito.")
            # Reiniciar la aplicacion para cargar los nuevos archivos (opcional)
        except subprocess.CalledProcessError as e:
            from app.core.logging import logger
            logger.error(f"Error generando Cerebro RAG: {e.stderr}", exc_info=True)
            
    background_tasks.add_task(run_generation)
    
    return {
        "status": "success", 
        "message": "Actualización del Cerebro RAG iniciada en segundo plano. Esto tomará de 5 a 15 minutos."
    }


@router.get("/materials/{material_id}/apus")
def get_material_apus(material_id: str, db: Session = Depends(get_db)):
    """Devuelve las partidas (APUs) donde se usa este material."""
    query = text(r'''
        SELECT a."CodPar", i."Descri", i."CovPar"
        FROM cost360_apu_materials a 
        JOIN cost360_items i ON a."CodPar" = i."CodPar" 
        WHERE a."CodIns" = :cod AND i."CovPar" ~ '^[A-Za-z]{1,2}[\.\-]?[0-9\.]+$'
    ''')
    rows = db.execute(query, {"cod": material_id}).fetchall()
    return [{"CodPar": r[0], "Descri": r[1], "CovPar": r[2]} for r in rows]

@router.get("/equipments/{equipment_id}/apus")
def get_equipment_apus(equipment_id: str, db: Session = Depends(get_db)):
    """Devuelve las partidas (APUs) donde se usa este equipo."""
    query = text(r'''
        SELECT a."CodPar", i."Descri", i."CovPar"
        FROM cost360_apu_equipment a 
        JOIN cost360_items i ON a."CodPar" = i."CodPar" 
        WHERE a."CodIns" = :cod AND i."CovPar" ~ '^[A-Za-z]{1,2}[\.\-]?[0-9\.]+$'
    ''')
    rows = db.execute(query, {"cod": equipment_id}).fetchall()
    return [{"CodPar": r[0], "Descri": r[1], "CovPar": r[2]} for r in rows]

@router.get("/labors/{labor_id}/apus")
def get_labor_apus(labor_id: str, db: Session = Depends(get_db)):
    """Devuelve las partidas (APUs) donde se usa esta mano de obra."""
    query = text(r'''
        SELECT a."CodPar", i."Descri", i."CovPar"
        FROM cost360_apu_labor a 
        JOIN cost360_items i ON a."CodPar" = i."CodPar" 
        WHERE a."CodIns" = :cod AND i."CovPar" ~ '^[A-Za-z]{1,2}[\.\-]?[0-9\.]+$'
    ''')
    rows = db.execute(query, {"cod": labor_id}).fetchall()
    return [{"CodPar": r[0], "Descri": r[1], "CovPar": r[2]} for r in rows]

