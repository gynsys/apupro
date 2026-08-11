"""
Nuevo endpoint de exportación de APU basado en apu_formulas.py
Usa notación de celdas (B1, H5) en lugar de row/col numéricos
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

def style_cell(cell, bold=False, size=11, align="left", border=False, number_format=None):
    cell.font = Font(bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        thin = Side(style='thin')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if number_format:
        cell.number_format = number_format

def generate_apu_excel(item, mat_rows, eq_rows, mo_rows):
    """Genera Excel con formato de apu_formulas.py"""
    rendimiento = item.RenPar or 1.0
    admin_gg = 16.0
    imprevisto_ut = 10.0
    financiamiento = 0.0
    iva = 0.0
    otros_imp = 0.0
    prestaciones = 435.0

    wb = Workbook()
    ws = wb.active
    ws.title = f"APU_{item.CovPar or item.CodPar}"

    # HEADER
    ws.merge_cells("B1:H1")
    ws["B1"] = "ANÁLISIS DE PRECIO UNITARIO"
    style_cell(ws["B1"], bold=True, size=14, align="center")
    
    ws.merge_cells("B2:H2")
    ws["B2"] = "ANÁLISIS DE PRECIO UNITARIO"
    style_cell(ws["B2"], bold=True, size=12, align="center")
    
    ws["B3"] = f"Obra: {item.Descri or 'N/A'}"
    ws["B4"] = f"Contratante: N/A"
    ws["E5"] = "Part. No.:"
    ws["F5"] = "1"
    ws["G5"] = "Fecha:"
    ws["H5"] = "46244"
    ws["B6"] = "Descripción:"
    ws.merge_cells("C6:H6")
    ws["C6"] = item.Descri or 'N/A'
    ws["G8"] = "Rendimiento:"
    ws["H8"] = rendimiento
    ws["B9"] = "Código:"
    ws["C9"] = item.CovPar or item.CodPar
    ws["E9"] = "Unidad:"
    ws["F9"] = item.UniPar
    ws["G9"] = "Cantidad:"
    ws["H9"] = "1"

    # MATERIALES
    mat_start = 11
    ws.merge_cells(f"B{mat_start}:H{mat_start}")
    ws[f"B{mat_start}"] = "MATERIALES"
    style_cell(ws[f"B{mat_start}"], bold=True, size=12)
    
    headers = ["No.", "Descripción", "Und.", "Cant.", "Desp.", "Precio", "Total"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 2)
        ws[f"{col}{mat_start+1}"] = h
        style_cell(ws[f"{col}{mat_start+1}"], bold=True, border=True)
    
    row = mat_start + 2
    for i, (apu_mat, mat) in enumerate(mat_rows):
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = mat.Descri if mat else ''
        ws[f"D{row}"] = mat.UniMat if mat else ''
        ws[f"E{row}"] = apu_mat.CanIns or 0
        ws[f"F{row}"] = apu_mat.Desper or 0
        ws[f"G{row}"] = mat.CosMat if mat else 0
        ws[f"H{row}"] = f"=ROUND((G{row}*E{row})*((F{row}/100)+1),2)"
        style_cell(ws[f"H{row}"], number_format='#,##0.00')
        row += 1
    
    total_mat_row = row
    ws[f"F{total_mat_row}"] = "Total Materiales:"
    first_data = mat_start + 2
    last_data = total_mat_row - 1
    ws[f"H{total_mat_row}"] = f"=SUM(H{first_data}:H{last_data})"
    style_cell(ws[f"H{total_mat_row}"], bold=True, number_format='#,##0.00')

    # EQUIPOS
    eq_start = total_mat_row + 2
    ws.merge_cells(f"B{eq_start}:H{eq_start}")
    ws[f"B{eq_start}"] = "EQUIPOS"
    style_cell(ws[f"B{eq_start}"], bold=True, size=12)
    
    headers = ["No.", "Descripción", "", "Cant.", "Cop/Dep", "Precio", "Total"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 2)
        ws[f"{col}{eq_start+1}"] = h
        style_cell(ws[f"{col}{eq_start+1}"], bold=True, border=True)
    
    row = eq_start + 2
    for i, (apu_eq, eq) in enumerate(eq_rows):
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = eq.Descri if eq else ''
        ws[f"E{row}"] = apu_eq.CanIns or 0
        ws[f"F{row}"] = apu_eq.Deprec or 0
        ws[f"G{row}"] = eq.CosDia if eq else 0
        ws[f"H{row}"] = f"=ROUND((G{row}*E{row})*(F{row}),2)"
        style_cell(ws[f"H{row}"], number_format='#,##0.00')
        row += 1
    
    total_eq_row = row
    ws[f"F{total_eq_row}"] = "Total Equipos:"
    first_data = eq_start + 2
    last_data = total_eq_row - 1
    ws[f"H{total_eq_row}"] = f"=SUM(H{first_data}:H{last_data})"
    style_cell(ws[f"H{total_eq_row}"], bold=True, number_format='#,##0.00')
    
    cuo_row = total_eq_row + 1
    ws[f"E{cuo_row}"] = "Costo Unitarios Equipos:"
    ws[f"H{cuo_row}"] = f"=ROUND(H{total_eq_row}/H9,2)"
    style_cell(ws[f"H{cuo_row}"], bold=True, number_format='#,##0.00')

    # MANO DE OBRA
    mo_start = cuo_row + 2
    ws.merge_cells(f"B{mo_start}:H{mo_start}")
    ws[f"B{mo_start}"] = "MANO DE OBRA"
    style_cell(ws[f"B{mo_start}"], bold=True, size=12)
    
    headers = ["No.", "Descripción", "Cant.", "Jornal", "Bono", "Total Jornal", "Total Bono"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 2)
        ws[f"{col}{mo_start+1}"] = h
        style_cell(ws[f"{col}{mo_start+1}"], bold=True, border=True)
    
    row = mo_start + 2
    for i, (apu_mo, mo) in enumerate(mo_rows):
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = mo.Descri if mo else ''
        ws[f"D{row}"] = apu_mo.CanIns or 0
        ws[f"E{row}"] = mo.Jornal if mo else 0
        ws[f"F{row}"] = mo.Bono if mo else 0
        ws[f"G{row}"] = f"=ROUND((D{row}*E{row}),2)"
        ws[f"H{row}"] = f"=ROUND((D{row}*F{row}),2)"
        style_cell(ws[f"G{row}"], number_format='#,##0.00')
        style_cell(ws[f"H{row}"], number_format='#,##0.00')
        row += 1
    
    sub_row = row
    first_data = mo_start + 2
    last_data = sub_row - 1
    ws[f"D{sub_row}"] = "SubTotal Mano de Obra:"
    ws[f"G{sub_row}"] = f"=SUM(G{first_data}:G{last_data})"
    ws[f"H{sub_row}"] = f"=SUM(H{first_data}:H{last_data})"
    style_cell(ws[f"G{sub_row}"], bold=True, number_format='#,##0.00')
    style_cell(ws[f"H{sub_row}"], bold=True, number_format='#,##0.00')
    
    ps_row = sub_row + 1
    ws[f"C{ps_row}"] = f"{prestaciones},00"
    ws[f"D{ps_row}"] = "Prestaciones Sociales:"
    ws[f"G{ps_row}"] = f"=ROUND((C{ps_row}/100)*G{sub_row},2)"
    ws[f"H{ps_row}"] = 0
    style_cell(ws[f"G{ps_row}"], number_format='#,##0.00')
    
    tg_row = ps_row + 1
    ws[f"D{tg_row}"] = "Total General Mano de Obra:"
    ws[f"H{tg_row}"] = f"=G{ps_row}+H{ps_row}+G{sub_row}+H{sub_row}"
    style_cell(ws[f"H{tg_row}"], bold=True, number_format='#,##0.00')
    
    cuo_mo_row = tg_row + 1
    ws[f"D{cuo_mo_row}"] = "Costo Unitario de Mano de Obra:"
    ws[f"H{cuo_mo_row}"] = f"=ROUND(H{tg_row}/H9,2)"
    style_cell(ws[f"H{cuo_mo_row}"], bold=True, number_format='#,##0.00')

    # RESUMEN
    resumen_start = cuo_mo_row + 2
    cd_row = resumen_start + 1
    ws[f"E{cd_row}"] = "COSTO DIRECTO SUBTOTAL A:"
    ws[f"H{cd_row}"] = f"=ROUND(H{total_mat_row}+H{total_eq_row}+H{cuo_mo_row},2)"
    style_cell(ws[f"H{cd_row}"], bold=True, number_format='#,##0.00')
    
    ad_row = cd_row + 1
    ws[f"C{ad_row}"] = f"{admin_gg},00"
    ws[f"D{ad_row}"] = "Administración y Gastos Generales:"
    ws[f"H{ad_row}"] = f"=ROUND((H{cd_row}*C{ad_row})/100,2)"
    style_cell(ws[f"H{ad_row}"], number_format='#,##0.00')
    
    sb_row = ad_row + 1
    ws[f"D{sb_row}"] = "SUBTOTAL B:"
    ws[f"H{sb_row}"] = f"=H{cd_row}+H{ad_row}"
    style_cell(ws[f"H{sb_row}"], bold=True, number_format='#,##0.00')
    
    iu_row = sb_row + 1
    ws[f"B{iu_row}"] = "SON: ( CATORCE MIL CIENTO CUARENTA Y UN Bs. con 78/100 ctms)"
    ws[f"E{iu_row}"] = f"{imprevisto_ut},00"
    ws[f"F{iu_row}"] = "Imprevisto Utilidad:"
    ws[f"H{iu_row}"] = f"=ROUND((H{sb_row}*E{iu_row})/100,2)"
    style_cell(ws[f"H{iu_row}"], number_format='#,##0.00')
    
    sc_row = iu_row + 1
    ws[f"D{sc_row}"] = "SUBTOTAL C:"
    ws[f"H{sc_row}"] = f"=H{sb_row}+H{iu_row}"
    style_cell(ws[f"H{sc_row}"], bold=True, number_format='#,##0.00')
    
    fin_row = sc_row + 1
    ws[f"E{fin_row}"] = f"{financiamiento},00"
    ws[f"F{fin_row}"] = "Financiamiento:"
    ws[f"H{fin_row}"] = f"=ROUND((H{sc_row}*E{fin_row})/100,2)"
    style_cell(ws[f"H{fin_row}"], number_format='#,##0.00')
    
    ps_row = fin_row + 1
    ws[f"D{ps_row}"] = "PRECIO UNITARIO SIN IMPUESTO:"
    ws[f"H{ps_row}"] = f"=H{sc_row}+H{fin_row}"
    style_cell(ws[f"H{ps_row}"], bold=True, number_format='#,##0.00')
    
    iva_row = ps_row + 1
    ws[f"E{iva_row}"] = f"{iva},00"
    ws[f"F{iva_row}"] = "Impuesto (I.V.A.):"
    ws[f"H{iva_row}"] = f"=ROUND((H{ps_row}*E{iva_row})/100,2)"
    style_cell(ws[f"H{iva_row}"], number_format='#,##0.00')
    
    oi_row = iva_row + 1
    ws[f"E{oi_row}"] = f"{otros_imp},00"
    ws[f"F{oi_row}"] = "Otros Impuestos:"
    ws[f"H{oi_row}"] = f"=ROUND((H{ps_row}*E{oi_row})/100,2)"
    style_cell(ws[f"H{oi_row}"], number_format='#,##0.00')
    
    pf_row = oi_row + 2
    ws[f"D{pf_row}"] = "PRECIO UNITARIO (Bs.F.):"
    ws[f"H{pf_row}"] = f"=H{ps_row}+H{iva_row}+H{oi_row}"
    style_cell(ws[f"H{pf_row}"], bold=True, number_format='#,##0.00')

    return wb
