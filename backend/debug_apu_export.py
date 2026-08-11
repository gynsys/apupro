#!/usr/bin/env python3
"""
Script de diagnóstico para exportación de APU a Excel
Prueba cada paso individualmente para identificar el problema exacto
"""

import sys
import os
sys.path.insert(0, '/app')

def test_openpyxl_import():
    """Prueba 1: Verificar que openpyxl está instalado"""
    print("=" * 60)
    print("TEST 1: Importar openpyxl")
    print("=" * 60)
    try:
        import openpyxl
        print(f"✅ openpyxl importado correctamente (versión {openpyxl.__version__})")
        return True
    except ImportError as e:
        print(f"❌ Error importando openpyxl: {e}")
        return False

def test_workbook_creation():
    """Prueba 2: Crear un workbook básico"""
    print("\n" + "=" * 60)
    print("TEST 2: Crear workbook básico")
    print("=" * 60)
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test"
        print("✅ Workbook creado exitosamente")
        return True
    except Exception as e:
        print(f"❌ Error creando workbook: {e}")
        return False

def test_database_connection():
    """Prueba 3: Conectar a la base de datos"""
    print("\n" + "=" * 60)
    print("TEST 3: Conectar a base de datos")
    print("=" * 60)
    try:
        from app.db.base import SessionLocal
        db = SessionLocal()
        print("✅ Conexión a base de datos exitosa")
        db.close()
        return True
    except Exception as e:
        print(f"❌ Error conectando a base de datos: {e}")
        return False

def test_get_item():
    """Prueba 4: Obtener un item de prueba"""
    print("\n" + "=" * 60)
    print("TEST 4: Obtener item CAR002")
    print("=" * 60)
    try:
        from app.db.base import SessionLocal
        from app.crud.crud_cost360 import get_item_by_code
        
        db = SessionLocal()
        item = get_item_by_code(db, "CAR002")
        print(f"✅ Item obtenido: {item.CodPar} - {item.Descri}")
        db.close()
        return True, item
    except Exception as e:
        print(f"❌ Error obteniendo item: {e}")
        return False, None

def test_get_apu_data(item_code):
    """Prueba 5: Obtener datos del APU"""
    print("\n" + "=" * 60)
    print("TEST 5: Obtener datos del APU")
    print("=" * 60)
    try:
        from app.db.base import SessionLocal
        from app.crud.crud_cost360 import get_apu_materials, get_apu_equipments, get_apu_labors
        
        db = SessionLocal()
        mat_rows = get_apu_materials(db, item_code)
        eq_rows = get_apu_equipments(db, item_code)
        mo_rows = get_apu_labors(db, item_code)
        
        print(f"✅ Materiales: {len(mat_rows)}")
        print(f"✅ Equipos: {len(eq_rows)}")
        print(f"✅ Mano de obra: {len(mo_rows)}")
        
        if mat_rows:
            print(f"   Primer material: {mat_rows[0]}")
            print(f"   Tipo: {type(mat_rows[0])}")
            if hasattr(mat_rows[0], '__len__'):
                print(f"   Longitud: {len(mat_rows[0])}")
        
        db.close()
        return True, mat_rows, eq_rows, mo_rows
    except Exception as e:
        print(f"❌ Error obteniendo datos del APU: {e}")
        import traceback
        traceback.print_exc()
        return False, None, None, None

def test_create_excel_with_data(mat_rows, eq_rows, mo_rows):
    """Prueba 6: Crear Excel con datos reales"""
    print("\n" + "=" * 60)
    print("TEST 6: Crear Excel con datos reales")
    print("=" * 60)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "APU Test"
        
        # Header
        ws['A1'] = "ANÁLISIS DE PRECIO UNITARIO"
        
        # Materiales
        ws['A3'] = "1. MATERIALES"
        ws['A4'] = "No."
        ws['B4'] = "Descripción"
        ws['C4'] = "Und."
        ws['D4'] = "Cant."
        ws['E4'] = "Desp."
        ws['F4'] = "Precio"
        ws['G4'] = "Total"
        
        row = 5
        for i, (apu_mat, mat) in enumerate(mat_rows):
            ws.cell(row, 1, i + 1)
            ws.cell(row, 2, mat.Descri if mat else '')
            ws.cell(row, 3, mat.UniMat if mat else '')
            ws.cell(row, 4, apu_mat.CanIns or 0)
            ws.cell(row, 5, apu_mat.Desper or 0)
            ws.cell(row, 6, mat.CosMat if mat else 0)
            row += 1
        
        # Guardar
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            print(f"✅ Excel creado exitosamente: {f.name}")
            return True, f.name
    except Exception as e:
        print(f"❌ Error creando Excel con datos: {e}")
        import traceback
        traceback.print_exc()
        return False, None

def test_with_formulas(mat_rows):
    """Prueba 7: Crear Excel con fórmulas"""
    print("\n" + "=" * 60)
    print("TEST 7: Crear Excel con fórmulas")
    print("=" * 60)
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Datos simples con fórmula
        ws['A1'] = 10
        ws['B1'] = 20
        ws['C1'] = "=A1*B1"
        
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            print(f"✅ Excel con fórmulas creado: {f.name}")
            return True, f.name
    except Exception as e:
        print(f"❌ Error creando Excel con fórmulas: {e}")
        import traceback
        traceback.print_exc()
        return False, None

def main():
    print("\n" + "=" * 60)
    print("DIAGNÓSTICO DE EXPORTACIÓN DE APU A EXCEL")
    print("=" * 60)
    
    results = []
    
    # Test 1: Import openpyxl
    results.append(("Import openpyxl", test_openpyxl_import()))
    if not results[-1][1]:
        print("\n❌ STOP: openpyxl no está instalado")
        return
    
    # Test 2: Crear workbook básico
    results.append(("Crear workbook", test_workbook_creation()))
    if not results[-1][1]:
        print("\n❌ STOP: No se puede crear workbook")
        return
    
    # Test 3: Conexión DB
    results.append(("Conexión DB", test_database_connection()))
    if not results[-1][1]:
        print("\n❌ STOP: No se puede conectar a DB")
        return
    
    # Test 4: Obtener item
    success, item = test_get_item()
    results.append(("Obtener item", success))
    if not success:
        print("\n❌ STOP: No se puede obtener item")
        return
    
    # Test 5: Obtener datos APU
    success, mat_rows, eq_rows, mo_rows = test_get_apu_data("CAR002")
    results.append(("Obtener datos APU", success))
    if not success:
        print("\n❌ STOP: No se pueden obtener datos del APU")
        return
    
    # Test 6: Crear Excel con datos
    success, file_path = test_create_excel_with_data(mat_rows, eq_rows, mo_rows)
    results.append(("Crear Excel con datos", success))
    
    # Test 7: Crear Excel con fórmulas
    success, formula_path = test_with_formulas(mat_rows)
    results.append(("Crear Excel con fórmulas", success))
    
    # Resumen
    print("\n" + "=" * 60)
    print("RESUMEN DE PRUEBAS")
    print("=" * 60)
    for test_name, success in results:
        status = "✅" if success else "❌"
        print(f"{status} {test_name}")
    
    print("\n" + "=" * 60)
    print("ARCHIVOS GENERADOS (si existen):")
    print("=" * 60)
    if file_path:
        print(f"Excel con datos: {file_path}")
    if formula_path:
        print(f"Excel con fórmulas: {formula_path}")

if __name__ == "__main__":
    main()
