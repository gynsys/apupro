import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def main():
    print("="*80)
    print("DETERMINACIÓN DE MAT A PROTEGER (POR DIFERENCIA)")
    print("="*80)
    
    # ==========================================
    # 1. LEER ARCHIVOS
    # ==========================================
    print("\n1. Leyendo archivos Excel...")
    
    archivo_original = "matches_para_revision_20260824_191225.xlsx"
    archivo_mod = "matches_para_revision_20260824_191225_MOD.xlsx"
    
    try:
        df_original = pd.read_excel(archivo_original)
        df_mod = pd.read_excel(archivo_mod)
        
        print(f"   Original: {len(df_original)} filas")
        print(f"   MOD: {len(df_mod)} filas")
        print(f"   Diferencia: {len(df_original) - len(df_mod)} filas eliminadas")
        
    except Exception as e:
        print(f"   Error leyendo archivos: {e}")
        return
    
    # ==========================================
    # 2. DETERMINAR MAT ELIMINADOS (A PROTEGER)
    # ==========================================
    print("\n2. Determinando MAT eliminados (a proteger)...")
    
    # Obtener códigos MAT de ambos archivos
    codigos_original = set(df_original.iloc[:, 0].dropna().astype(str))  # Columna 0: Código MAT
    codigos_mod = set(df_mod.iloc[:, 0].dropna().astype(str))
    
    # Diferencia = los que están en original pero NO en mod
    codigos_proteger = codigos_original - codigos_mod
    
    print(f"   MAT a proteger con 'R': {len(codigos_proteger)}")
    
    if len(codigos_proteger) > 0:
        print(f"   Ejemplos: {list(codigos_proteger)[:5]}")
    
    # ==========================================
    # 3. MARCAR MAT PROTEGIDOS EN ARCHIVO ORIGINAL
    # ==========================================
    print("\n3. Marcando MAT protegidos con 'R'...")
    
    # Copia del archivo original para modificar
    df_final = df_original.copy()
    
    # Crear columna de protección
    df_final['Proteger'] = ''
    
    # Marcar los MAT a proteger con "R"
    for idx, row in df_final.iterrows():
        cod_mat = str(row.iloc[0])  # Código MAT (columna 0)
        if cod_mat in codigos_proteger:
            df_final.at[idx, 'Proteger'] = 'R'
    
    protegidos_count = (df_final['Proteger'] == 'R').sum()
    print(f"   MAT marcados con 'R': {protegidos_count}")
    
    # ==========================================
    # 4. GENERAR EXCEL CON FORMATO
    # ==========================================
    print("\n4. Generando Excel con MAT protegidos identificados...")
    
    wb = load_workbook(archivo_original)
    ws = wb.active
    
    # Añadir columna de protección
    col_proteger = len(df_final.columns) + 1
    ws.cell(row=1, column=col_proteger, value="Proteger")
    
    # Formato para header
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for row in range(1, ws.max_row + 1):
        cod_mat = ws.cell(row=row, column=1).value  # Columna 1: Código MAT
        
        if row == 1:  # Header
            cell = ws.cell(row=row, column=col_proteger)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            if cod_mat and str(cod_mat) in codigos_proteger:
                # Marcar con "R" y color rojo
                cell = ws.cell(row=row, column=col_proteger, value="R")
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar ancho de columna
    ws.column_dimensions[get_column_letter(col_proteger)].width = 10
    
    # Guardar archivo
    filename_final = f"matches_para_revision_PROTEGIDOS_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename_final)
    
    print(f"   Excel generado: {filename_final}")
    
    # ==========================================
    # 5. REPORTE DE MAT PROTEGIDOS
    # ==========================================
    print("\n5. Lista de MAT protegidos (marcados con 'R'):")
    
    mat_protegidos_info = []
    for cod in sorted(codigos_proteger):
        # Buscar información del MAT original
        fila_original = df_original[df_original.iloc[:, 0].astype(str) == cod]
        if not fila_original.empty:
            info = {
                'codigo': cod,
                'descripcion': fila_original.iloc[0, 1] if len(fila_original.columns) > 1 else 'N/A',
                'precio': fila_original.iloc[0, 2] if len(fila_original.columns) > 2 else 'N/A'
            }
            mat_protegidos_info.append(info)
            print(f"   - {cod}: {info['descripcion'][:50]}... ${info['precio']}")
    
    # ==========================================
    # 6. GUARDAR REPORTE SEPARADO
    # ==========================================
    print("\n6. Generando reporte de MAT protegidos...")
    
    if mat_protegidos_info:
        df_protegidos = pd.DataFrame(mat_protegidos_info)
        filename_reporte = f"mat_protegidos_R_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        df_protegidos.to_excel(filename_reporte, index=False)
        print(f"   Reporte guardado: {filename_reporte}")
    
    print("\n" + "="*80)
    print("PROCESO COMPLETADO")
    print("="*80)
    print(f"\nTotal MAT protegidos: {len(codigos_proteger)}")
    print(f"Excel final: {filename_final}")
    print(f"Reporte MAT protegidos: {filename_reporte if mat_protegidos_info else 'N/A'}")
    print("\nEstos MAT están marcados con 'R' y serán protegidos en la fusión.")

if __name__ == '__main__':
    main()