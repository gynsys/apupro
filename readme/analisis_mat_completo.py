import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import re
from difflib import SequenceMatcher

def normalizar_descripcion(texto):
    """Normaliza texto para comparación mejorada con reglas específicas"""
    if not texto:
        return ""
    
    # Convertir a mayúsculas para procesamiento (luego minúsculas)
    texto = texto.upper()
    
    # REGLAS ESPECÍFICAS para casos emblemáticos:
    
    # 1. Eliminar puntos entre letras (H.G. -> HG, M.M -> MM, W.C -> WC)
    texto = re.sub(r'([A-Z])\.([A-Z])', r'\1\2', texto)  # Puntos entre letras
    texto = re.sub(r'([A-Z])\.([A-Z])', r'\1\2', texto)  # Repetir para casos múltiples
    
    # 2. Normalizar unidades (KGF. -> KG, KG -> KG, LT -> LITRO, etc.)
    texto = re.sub(r'KGF\.?', 'KG', texto)
    texto = re.sub(r'KG\.?', 'KG', texto)
    texto = re.sub(r'LTS?', 'LITRO', texto)
    texto = re.sub(r'MTS?', 'METRO', texto)
    texto = re.sub(r'CMS?', 'CENTIMETRO', texto)
    texto = re.sub(r'MMS?', 'MILIMETRO', texto)
    texto = re.sub(r'PZS?', 'PIEZA', texto)
    texto = re.sub(r'UNDS?', 'UNIDAD', texto)
    
    # 3. Eliminar guiones y reemplazar por espacios
    texto = re.sub(r'-', ' ', texto)
    
    # 4. Normalizar números (24.000 -> 24000, pero mantener decimales 1.5 -> 1.5)
    # Solo eliminar puntos si son separadores de miles (seguidos de 3+ dígitos)
    texto = re.sub(r'(\d+)\.(\d{3,})', r'\1\2', texto)
    
    # 5. Eliminar espacios alrededor de puntuación restante
    texto = re.sub(r'\s+', ' ', texto)
    
    # Convertir a minúsculas para comparación
    texto = texto.lower()
    
    # 6. Eliminar puntuación restante excepto espacios
    texto = re.sub(r'[^\w\s]', ' ', texto)
    
    # 7. Eliminar palabras comunes que no agregan valor
    stop_words = ['con', 'sin', 'para', 'de', 'la', 'el', 'los', 'las', 'un', 'una', 'y', 'o', 'en', 'c/', 'p/', 'c', 'p']
    words = texto.split()
    words = [w for w in words if w not in stop_words and len(w) > 2]
    
    return ' '.join(sorted(words))  # Ordenar palabras para match independiente del orden

def calcular_similitud(texto1, texto2):
    """Calcula similitud entre dos textos usando SequenceMatcher"""
    return SequenceMatcher(None, texto1, texto2).ratio()

def extraer_palabras_clave(texto):
    """Extrae palabras clave significativas"""
    if not texto:
        return []
    
    texto = texto.lower()
    texto = re.sub(r'[^\w\s]', ' ', texto)
    words = texto.split()
    
    # Filtrar palabras cortas y comunes
    stop_words = ['con', 'sin', 'para', 'de', 'la', 'el', 'los', 'las', 'un', 'una', 'y', 'o', 'en', 'c/', 'p/']
    keywords = [w for w in words if len(w) > 3 and w not in stop_words]
    
    return keywords

def main():
    print("Conectando a la base de datos...")
    engine = create_engine('postgresql://apupro_user:apupro_password@costbase.net:5440/apupro_db')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis MAT"
    
    # ==========================================
    # CONFIGURACIÓN DE FORMATOS EXCEL
    # ==========================================
    
    # Colores
    header_fill = PatternFill(start_color="1A6BB5", end_color="1A6BB5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    match_exacto_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
    match_parcial_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
    sin_match_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Rosa claro
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==========================================
    # ENCABEZADOS
    # ==========================================
    
    headers = [
        "Código MAT",
        "Descripción MAT", 
        "Precio MAT",
        "Match",
        "Código Match",
        "Descripción Match",
        "Precio Match",
        "Score %",
        "Diferencia $",
        "Diferencia %",
        "Tipo Match",
        "Propuesta Fusión",
        "Código Propuesto",
        "Observaciones"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Ajustar ancho de columnas
    column_widths = [15, 40, 12, 10, 15, 40, 12, 10, 12, 10, 15, 15, 15, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # ==========================================
    # ANÁLISIS DE MATERIALES MAT
    # ==========================================
    
    with engine.connect() as conn:
        print("Obteniendo todos los materiales MAT y MATR...")
        todos_mat = conn.execute(text('''
            SELECT "CodMat", "Descri", "CosMat"
            FROM cost360_materials 
            WHERE "CodMat" LIKE 'MAT%' OR "CodMat" LIKE 'MATR%'
            ORDER BY "CodMat"
        ''')).fetchall()
        
        print(f"Total MAT/MATR a analizar: {len(todos_mat)}")
        
        # Obtener todos los materiales no-MAT y no-MATR para matching
        print("Obteniendo materiales correctos (no-MAT, no-MATR)...")
        todos_no_mat = conn.execute(text('''
            SELECT "CodMat", "Descri", "CosMat"
            FROM cost360_materials 
            WHERE "CodMat" NOT LIKE 'MAT%' AND "CodMat" NOT LIKE 'MATR%'
            ORDER BY "CodMat"
        ''')).fetchall()
        
        print(f"Total materiales correctos: {len(todos_no_mat)}")
        
        # Crear diccionario con normalización para búsqueda mejorada
        # Y crear índice por palabras clave para búsquedas más rápidas
        materiales_dict = {}
        keywords_index = {}  # {keyword: [cod1, cod2, ...]}
        
        for cod, descri, precio in todos_no_mat:
            descri_norm = normalizar_descripcion(descri)
            keywords = extraer_palabras_clave(descri)
            
            materiales_dict[cod] = {
                'descri': descri,
                'precio': precio,
                'descri_lower': descri.lower(),
                'descri_norm': descri_norm,
                'keywords': keywords
            }
            
            # Indexar por palabras clave
            for kw in keywords:
                if kw not in keywords_index:
                    keywords_index[kw] = []
                keywords_index[kw].append(cod)
        
        row_num = 2
        resultados = {
            'match_exacto': 0,
            'match_parcial': 0,
            'sin_match': 0
        }
        
        print("Procesando coincidencias...")
        
        for mat_cod, mat_descri, mat_precio in todos_mat:
            mat_descri_lower = mat_descri.lower().strip()
            mat_descri_norm = normalizar_descripcion(mat_descri)
            mat_keywords = extraer_palabras_clave(mat_descri)
            
            # Buscar match exacto (con normalización)
            match_exacto = None
            for cod, datos in materiales_dict.items():
                if datos['descri_norm'] == mat_descri_norm:
                    match_exacto = (cod, datos['descri'], datos['precio'])
                    break
            
            # Si no hay match exacto, buscar por palabras clave (más eficiente)
            match_keywords = None
            if not match_exacto and len(mat_keywords) >= 2:
                # Encontrar candidatos usando índice de keywords
                candidatos = set()
                for kw in mat_keywords:
                    if kw in keywords_index:
                        candidatos.update(keywords_index[kw])
                
                # Evaluar solo los candidatos
                mejor_kw_match = 0
                for cod in candidatos:
                    keywords_match = sum(1 for kw in mat_keywords if kw in materiales_dict[cod]['keywords'])
                    if keywords_match > mejor_kw_match and keywords_match >= len(mat_keywords) * 0.7:
                        match_keywords = (cod, materiales_dict[cod]['descri'], materiales_dict[cod]['precio'], keywords_match)
                        mejor_kw_match = keywords_match
            
            # Si no hay keywords match, buscar fuzzy match solo sobre candidatos
            match_fuzzy = None
            mejor_score = 0
            if not match_exacto and not match_keywords and len(mat_keywords) >= 1:
                # Usar keywords para encontrar candidatos para fuzzy matching
                candidatos = set()
                for kw in mat_keywords:
                    if kw in keywords_index:
                        candidatos.update(keywords_index[kw])
                
                # Si no hay candidatos por keywords, usar muestra aleatoria
                if not candidatos:
                    import random
                    sample_keys = random.sample(list(materiales_dict.keys()), min(100, len(materiales_dict)))
                    candidatos = sample_keys
                
                # Evaluar fuzzy match solo sobre candidatos
                for cod in candidatos:
                    score = calcular_similitud(mat_descri_norm, materiales_dict[cod]['descri_norm'])
                    if score > 0.85 and score > mejor_score:  # Umbral más alto
                        match_fuzzy = (cod, materiales_dict[cod]['descri'], materiales_dict[cod]['precio'], score)
                        mejor_score = score
            
            # Determinar tipo de match y valores
            if match_exacto:
                tipo_match = "EXACTO"
                match_cod, match_descri, match_precio = match_exacto
                resultados['match_exacto'] += 1
                fill = match_exacto_fill
                propuesta = "FUSIONAR"
                cod_propuesto = match_cod
                obs = "Match exacto (normalizado)"
                
            elif match_fuzzy:
                tipo_match = "FUZZY"
                match_cod, match_descri, match_precio, score = match_fuzzy
                resultados['match_parcial'] += 1
                fill = match_parcial_fill
                propuesta = "REVISAR" if score < 0.9 else "FUSIONAR"
                cod_propuesto = match_cod
                obs = f"Fuzzy match {score:.1%} - requiere validación"
                
            elif match_keywords:
                tipo_match = "KEYWORDS"
                match_cod, match_descri, match_precio, kw_match = match_keywords
                resultados['match_parcial'] += 1
                fill = match_parcial_fill
                propuesta = "REVISAR"
                cod_propuesto = match_cod
                obs = f"Match por keywords ({kw_match}/{len(mat_keywords)})"
                
            else:
                tipo_match = "SIN MATCH"
                match_cod, match_descri, match_precio = "", "", 0
                resultados['sin_match'] += 1
                fill = sin_match_fill
                propuesta = "MANUAL"
                cod_propuesto = ""
                obs = "No se encontró coincidencia"
            
            # Calcular diferencias y score
            if match_precio > 0:
                diff_dolares = abs(mat_precio - match_precio)
                diff_porcentaje = ((match_precio - mat_precio) / mat_precio * 100) if mat_precio > 0 else 0
            else:
                diff_dolares = 0
                diff_porcentaje = 0
            
            # Obtener score si existe
            score_value = ""
            if match_fuzzy:
                score_value = f"{match_fuzzy[3]:.1%}"
            elif match_keywords:
                score_value = f"{match_keywords[3]}/{len(mat_keywords)}"
            
            # Escribir fila
            row_data = [
                mat_cod,
                mat_descri,
                round(mat_precio, 2),
                "SÍ" if match_exacto or match_fuzzy or match_keywords else "NO",
                match_cod,
                match_descri,
                round(match_precio, 2) if match_precio > 0 else "",
                score_value,
                round(diff_dolares, 2) if diff_dolares > 0 else "",
                f"{diff_porcentaje:.1f}%" if diff_porcentaje != 0 else "",
                tipo_match,
                propuesta,
                cod_propuesto,
                obs
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Formato numérico para precios
                if col_num in [3, 7, 9]:
                    if value and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
            
            row_num += 1
            
            # Mostrar progreso cada 100 registros
            if row_num % 100 == 0:
                print(f"Procesados {row_num - 1} de {len(todos_mat)} registros...")
    
    # ==========================================
    # RESUMEN FINAL
    # ==========================================
    
    print("\n" + "="*60)
    print("RESUMEN DEL ANÁLISIS")
    print("="*60)
    print(f"Total MAT analizados: {len(todos_mat)}")
    print(f"Match exacto: {resultados['match_exacto']} ({resultados['match_exacto']/len(todos_mat)*100:.1f}%)")
    print(f"Match parcial: {resultados['match_parcial']} ({resultados['match_parcial']/len(todos_mat)*100:.1f}%)")
    print(f"Sin match: {resultados['sin_match']} ({resultados['sin_match']/len(todos_mat)*100:.1f}%)")
    
    # ==========================================
    # GUARDAR ARCHIVO EXCEL
    # ==========================================
    
    filename = f"analisis_mat_completo_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    print(f"\nArchivo Excel generado: {filename}")
    print("El archivo contiene:")
    print("  - Columnas con formato y colores por tipo de match")
    print("  - Propuesta de fusión para cada material MAT")
    print("  - Código propuesto para reemplazo")
    print("  - Observaciones para validación manual")

if __name__ == '__main__':
    main()