import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def main():
    print("="*80)
    print("TRIPLE DIFERENCIA: ORIGINAL - MOD - PRIMERA_LISTA")
    print("="*80)
    
    # ==========================================
    # 1. LEER LOS TRES ARCHIVOS
    # ==========================================
    print("\n1. Leyendo archivos Excel...")
    
    archivo_original = "matches_para_revision_20260824_191225.xlsx"
    archivo_mod = "matches_para_revision_20260824_191225_MOD.xlsx"
    archivo_primera = "PRIMERA_LISTA_A_ FUSIONAR.xlsx"
    
    try:
        df_original = pd.read_excel(archivo_original)
        df_mod = pd.read_excel(archivo_mod)
        df_primera = pd.read_excel(archivo_primera)
        
        print(f"   Original: {len(df_original)} filas")
        print(f"   MOD: {len(df_mod)} filas")
        print(f"   PRIMERA_LISTA: {len(df_primera)} filas")
        
    except Exception as e:
        print(f"   Error leyendo archivos: {e}")
        return
    
    # ==========================================
    # 2. OBTENER CÓDIGOS MAT DE CADA ARCHIVO
    # ==========================================
    print("\n2. Extrayendo códigos MAT...")
    
    codigos_original = set(df_original.iloc[:, 0].dropna().astype(str))
    codigos_mod = set(df_mod.iloc[:, 0].dropna().astype(str))
    codigos_primera = set(df_primera.iloc[:, 0].dropna().astype(str))
    
    print(f"   Códigos en original: {len(codigos_original)}")
    print(f"   Códigos en MOD: {len(codigos_mod)}")
    print(f"   Códigos en PRIMERA_LISTA: {len(codigos_primera)}")
    
    # ==========================================
    # 3. REALIZAR TRIPLE DIFERENCIA
    # ==========================================
    print("\n3. Realizando triple diferencia...")
    print("   ORIGINAL - MOD - PRIMERA_LISTA")
    
    # Paso 1: Original - MOD = MAT que eliminaste (413)
    diff_1 = codigos_original - codigos_mod
    print(f"   Original - MOD: {len(diff_1)} códigos")
    
    # Paso 2: Diff_1 - PRIMERA_LISTA = MAT restantes
    diff_final = diff_1 - codigos_primera
    print(f"   (Original - MOD) - PRIMERA_LISTA: {len(diff_final)} códigos")
    
    # ==========================================
    # 4. OBTENER INFORMACIÓN DE LOS MAT RESULTANTES
    # ==========================================
    print("\n4. Obteniendo información de los MAT resultantes...")
    
    mat_resultantes_info = []
    for cod in sorted(diff_final):
        # Buscar información del archivo original
        fila_original = df_original[df_original.iloc[:, 0].astype(str) == cod]
        if not fila_original.empty:
            info = {
                'codigo': cod,
                'descripcion': fila_original.iloc[0, 1] if len(fila_original.columns) > 1 else 'N/A',
                'precio': fila_original.iloc[0, 2] if len(fila_original.columns) > 2 else 'N/A'
            }
            mat_resultantes_info.append(info)
    
    print(f"   MAT resultantes con información: {len(mat_resultantes_info)}")
    
    # ==========================================
    # 5. GENERAR EXCEL CON RESULTADOS
    # ==========================================
    print("\n5. Generando Excel con MAT resultantes...")
    
    if mat_resultantes_info:
        df_resultantes = pd.DataFrame(mat_resultantes_info)
        filename = f"mat_triple_diferencia_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        df_resultantes.to_excel(filename, index=False)
        print(f"   Excel generado: {filename}")
    
    # ==========================================
    # 6. MOSTRAR RESULTADOS
    # ==========================================
    print("\n" + "="*80)
    print("RESULTADOS DE LA TRIPLE DIFERENCIA")
    print("="*80)
    print(f"\nTotal MAT resultantes: {len(diff_final)}")
    print(f"\nResumen:")
    print(f"  - Original tenía: {len(codigos_original)} MAT")
    print(f"  - Eliminaste en MOD: {len(diff_1)} MAT")
    print(f"  - Ya en PRIMERA_LISTA: {len(codigos_primera)} MAT")
    print(f"  - Nuevos MAT disponibles: {len(diff_final)}")
    
    if mat_resultantes_info:
        print(f"\nEjemplos de MAT resultantes:")
        for i, info in enumerate(mat_resultantes_info[:20]):
            print(f"  {i+1}. {info['codigo']}: {info['descripcion'][:50]}... ${info['precio']}")
        
        if len(mat_resultantes_info) > 20:
            print(f"  ... y {len(mat_resultantes_info) - 20} más")
    
    print(f"\nArchivo con lista completa: {filename if mat_resultantes_info else 'N/A'}")

if __name__ == '__main__':
    main()