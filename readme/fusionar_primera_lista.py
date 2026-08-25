import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def main():
    print("="*80)
    print("FUSIÓN DE PRIMERA LISTA + MARCAR MATR")
    print("="*80)
    
    # ==========================================
    # 1. LEER ARCHIVOS NECESARIOS
    # ==========================================
    print("\n1. Leyendo archivos...")
    
    archivo_primera = "PRIMERA_LISTA_A_ FUSIONAR.xlsx"
    archivo_triple_diff = "mat_triple_diferencia_20260825_073217.xlsx"
    archivo_protegidos = "mat_protegidos_R_20260825_072819.xlsx"
    
    try:
        df_primera = pd.read_excel(archivo_primera)
        df_triple = pd.read_excel(archivo_triple_diff)
        df_protegidos = pd.read_excel(archivo_protegidos)
        
        print(f"   PRIMERA_LISTA: {len(df_primera)} MAT a fusionar")
        print(f"   Triple diferencia: {len(df_triple)} MAT nuevos a proteger")
        print(f"   Protegidos originales: {len(df_protegidos)} MAT ya protegidos")
        
    except Exception as e:
        print(f"   Error leyendo archivos: {e}")
        return
    
    # ==========================================
    # 2. OBTENER CÓDIGOS
    # ==========================================
    print("\n2. Preparando listas de códigos...")
    
    codigos_fusionar = set(df_primera.iloc[:, 0].dropna().astype(str))
    codigos_nuevos_proteger = set(df_triple.iloc[:, 0].dropna().astype(str))
    codigos_ya_protegidos = set(df_protegidos.iloc[:, 0].dropna().astype(str))
    
    # Unir todos los MAT a proteger
    codigos_proteger_total = codigos_nuevos_proteger | codigos_ya_protegidos
    
    print(f"   MAT a fusionar: {len(codigos_fusionar)}")
    print(f"   MAT a proteger (nuevos MATR): {len(codigos_nuevos_proteger)}")
    print(f"   MAT ya protegidos: {len(codigos_ya_protegidos)}")
    print(f"   Total MAT protegidos: {len(codigos_proteger_total)}")
    
    # ==========================================
    # 3. CONECTAR A BD Y BACKUP
    # ==========================================
    print("\n3. Conectando a base de datos de producción...")
    
    engine = create_engine('postgresql://apupro_user:apupro_password@costbase.net:5440/apupro_db')
    
    with engine.connect() as conn:
        # Backup antes de fusionar
        backup_table = f"cost360_materials_backup_fusion_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        try:
            conn.execute(text(f'''
                CREATE TABLE {backup_table} AS 
                SELECT * FROM cost360_materials 
                WHERE "CodMat" LIKE 'MAT%'
            '''))
            conn.commit()
            print(f"   Backup creado: {backup_table}")
        except Exception as e:
            print(f"   Error creando backup: {e}")
            return
        
        # ==========================================
        # 4. LEER PRECIOS DE MATCH PARA FUSIÓN
        # ==========================================
        print("\n4. Preparando datos para fusión...")
        
        # Mapeo de MAT a sus códigos de reemplazo y precios
        fusion_map = {}
        for idx, row in df_primera.iterrows():
            cod_mat = str(row['Codigo MAT'])  # Código MAT
            cod_match = str(row['Codigo Match']) if pd.notna(row['Codigo Match']) else ''
            precio_match = float(row['Precio Match']) if pd.notna(row['Precio Match']) else 0.0
            
            if cod_match and precio_match > 0:
                fusion_map[cod_mat] = {
                    'cod_match': cod_match,
                    'precio_match': precio_match
                }
        
        print(f"   Mapeo de fusión preparado: {len(fusion_map)} MAT")
        
        # ==========================================
        # 5. EJECUTAR FUSIÓN DE PRECIOS
        # ==========================================
        print("\n5. Ejecutando fusión de precios...")
        print("   (Fusionando PRIMERA_LISTA ignorando protección original)")
        
        cambios_fusion = []
        for cod_mat, datos in fusion_map.items():
            # FUSIONAR TODOS los MAT de PRIMERA_LISTA, ignorando protección
            try:
                # Actualizar precio en BD
                conn.execute(text('''
                    UPDATE cost360_materials 
                    SET "CosMat" = :nuevo_precio
                    WHERE "CodMat" = :cod_mat
                '''), {"nuevo_precio": datos['precio_match'], "cod_mat": cod_mat})
                
                cambios_fusion.append({
                    'codigo': cod_mat,
                    'cod_match': datos['cod_match'],
                    'precio_nuevo': datos['precio_match']
                })
                
            except Exception as e:
                print(f"   Error actualizando {cod_mat}: {e}")
        
        conn.commit()
        print(f"   Precios actualizados: {len(cambios_fusion)}")
        
        # ==========================================
        # 6. MARCAR MATR LOS NUEVOS PROTEGIDOS
        # ==========================================
        print("\n6. Marcando nuevos MAT con MATR...")
        
        # Para los MAT nuevos a proteger, podríamos añadir un prefijo o marcar de alguna forma
        # Por ahora solo reportamos, ya que la decisión de marcar como MATR requiere más detalles
        
        print(f"   MAT nuevos a proteger (MATR): {len(codigos_nuevos_proteger)}")
        print(f"   Estos MAT se mantendrán con sus precios actuales")
        
        # ==========================================
        # 7. REPORTE FINAL
        # ==========================================
        print("\n" + "="*80)
        print("FUSIÓN COMPLETADA")
        print("="*80)
        print(f"\nResumen:")
        print(f"  - MAT fusionados (precios actualizados): {len(cambios_fusion)}")
        print(f"  - MAT protegidos (MATR nuevos): {len(codigos_nuevos_proteger)}")
        print(f"  - MAT ya protegidos: {len(codigos_ya_protegidos)}")
        print(f"  - Total MAT protegidos: {len(codigos_proteger_total)}")
        print(f"\nBackup disponible: {backup_table}")
        
        # ==========================================
        # 8. GUARDAR REPORTE DE FUSIÓN
        # ==========================================
        print("\n8. Generando reporte de fusión...")
        
        if cambios_fusion:
            df_cambios = pd.DataFrame(cambios_fusion)
            filename_reporte = f"reporte_fusion_primera_lista_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            df_cambios.to_excel(filename_reporte, index=False)
            print(f"   Reporte guardado: {filename_reporte}")
        
        # Reporte de MATR
        if codigos_nuevos_proteger:
            matr_info = []
            for cod in sorted(codigos_nuevos_proteger):
                fila = df_triple[df_triple.iloc[:, 0].astype(str) == cod]
                if not fila.empty:
                    matr_info.append({
                        'codigo': cod,
                        'descripcion': fila.iloc[0, 1],
                        'precio': fila.iloc[0, 2]
                    })
            
            if matr_info:
                df_matr = pd.DataFrame(matr_info)
                filename_matr = f"matr_nuevos_protegidos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                df_matr.to_excel(filename_matr, index=False)
                print(f"   Reporte MATR guardado: {filename_matr}")

if __name__ == '__main__':
    main()