import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import re
from difflib import SequenceMatcher
from typing import Optional


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE DIÁMETRO
# ─────────────────────────────────────────────────────────────────────────────

def extraer_diametro(texto: str) -> Optional[str]:
    """
    Extrae y normaliza el diámetro de una descripción.
    Retorna None si no hay diámetro.
    Ejemplos:
      D=2"       → "2IN"
      D=1 1/2"   → "11/2IN"
      D=110MM    → "110MM"
      D=2.5"     → "2.5IN"
      D= 3/4"    → "3/4IN"
    """
    if not texto:
        return None

    t = texto.upper().strip()

    # Patrón: D= seguido de valor numérico con posible fracción y unidad
    match = re.search(
        r'\bD\s*=\s*'                          # D=
        r'(\d+(?:[.,]\d+)?'                    # parte entera (ej: 1, 2.5, 24)
        r'(?:\s+\d+/\d+)?'                     # fracción opcional (ej: 1/2, 3/4)
        r'|'                                   # ─── O ───
        r'\d+/\d+)'                            # solo fracción (ej: 3/4, 1/2)
        r'\s*'                                 # espacio opcional
        r'("|\'\'|MM|CM|PLG|PULG|M\b)?',      # unidad opcional
        t
    )

    if not match:
        return None

    valor = re.sub(r'\s+', '', match.group(1))   # quitar espacios internos: "1 1/2" → "11/2"
    unidad = (match.group(2) or '"').strip()      # default pulgadas si no hay unidad

    # Normalizar unidad
    unidad_norm = {
        '"': 'IN', "''": 'IN', 'PLG': 'IN', 'PULG': 'IN',
        'MM': 'MM', 'CM': 'CM', 'M': 'M'
    }.get(unidad, 'IN')

    return f"{valor}{unidad_norm}"


def diametros_compatibles(diam1: Optional[str], diam2: Optional[str]) -> bool:
    """
    True si los diámetros son compatibles para hacer match.
    - Si alguno no tiene diámetro → compatible (no aplica restricción)
    - Si ambos tienen diámetro → deben ser iguales
    """
    if diam1 is None or diam2 is None:
        return True
    return diam1 == diam2


# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZACIÓN DE DESCRIPCIÓN
# ─────────────────────────────────────────────────────────────────────────────

def normalizar_descripcion(texto: str) -> str:
    """Normaliza texto para comparación. NO incluye diámetro (se compara por separado)."""
    if not texto:
        return ""

    texto = texto.upper()

    # 1. Eliminar puntos entre letras (H.G. → HG)
    texto = re.sub(r'([A-Z])\.([A-Z])', r'\1\2', texto)
    texto = re.sub(r'([A-Z])\.([A-Z])', r'\1\2', texto)

    # 2. Normalizar unidades de medida
    texto = re.sub(r'KGF\.?', 'KG', texto)
    texto = re.sub(r'KG\.?', 'KG', texto)
    texto = re.sub(r'\bLTS?\b', 'LITRO', texto)
    texto = re.sub(r'\bMTS?\b', 'METRO', texto)
    texto = re.sub(r'\bCMS?\b', 'CENTIMETRO', texto)
    texto = re.sub(r'\bMMS?\b', 'MILIMETRO', texto)
    texto = re.sub(r'\bPZS?\b', 'PIEZA', texto)
    texto = re.sub(r'\bUNDS?\b', 'UNIDAD', texto)

    # 3. Separador de miles (18.000 → 18000)
    texto = re.sub(r'(\d{1,3})\.(\d{3})(?!\d)', r'\1\2', texto)

    # 4. Eliminar diámetro/dimensión de la normalización
    #    El diámetro se compara por separado vía extraer_diametro()
    texto = re.sub(r'\bD\s*=\s*[\d/\s.,]+(?:"|\'\'|MM|CM|PLG|PULG|M\b)?', '', texto)

    # 5. Eliminar guiones
    texto = re.sub(r'-', ' ', texto)

    # 6. Coma entre comilla y dimensión
    texto = re.sub(r'",\s*', '" ', texto)

    # 7. Limpiar espacios
    texto = re.sub(r'\s+', ' ', texto).strip()

    # 8. Pasar a minúsculas
    texto = texto.lower()

    # 9. Eliminar puntuación restante
    texto = re.sub(r'[^\w\s]', ' ', texto)

    # 10. Stop words y palabras cortas
    stop_words = {
        'con', 'sin', 'para', 'de', 'la', 'el', 'los', 'las',
        'un', 'una', 'y', 'o', 'en', 'tipo', 'por'
    }
    words = texto.split()
    words = [w for w in words if w not in stop_words and len(w) > 2]

    return ' '.join(sorted(words))


def calcular_similitud(texto1: str, texto2: str) -> float:
    """Calcula similitud entre dos textos usando SequenceMatcher."""
    return SequenceMatcher(None, texto1, texto2).ratio()


def extraer_palabras_clave(texto: str) -> list:
    """Extrae palabras clave significativas."""
    if not texto:
        return []
    texto = texto.lower()
    texto = re.sub(r'[^\w\s]', ' ', texto)
    stop_words = {
        'con', 'sin', 'para', 'de', 'la', 'el', 'los', 'las',
        'un', 'una', 'y', 'o', 'en', 'tipo'
    }
    return [w for w in texto.split() if len(w) > 3 and w not in stop_words]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Conectando a la base de datos...")
    engine = create_engine('postgresql://apupro_user:apupro_password@costbase.net:5440/apupro_db')

    wb = Workbook()
    ws = wb.active
    ws.title = "Analisis MAT"

    # Colores
    header_fill      = PatternFill(start_color="1A6BB5", end_color="1A6BB5", fill_type="solid")
    header_font      = Font(bold=True, color="FFFFFF", size=11)
    match_exacto_fill  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    match_parcial_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    sin_match_fill     = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    diam_block_fill    = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja: bloqueado por diámetro

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    headers = [
        "Codigo MAT", "Descripcion MAT", "Precio MAT",
        "Diametro MAT",
        "Match", "Codigo Match", "Descripcion Match", "Precio Match",
        "Diametro Match",
        "Score %", "Diferencia $", "Diferencia %",
        "Tipo Match", "Propuesta Fusion", "Codigo Propuesto", "Observaciones"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    column_widths = [14, 42, 11, 12, 6, 14, 42, 11, 12, 9, 11, 10, 13, 14, 14, 32]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    with engine.connect() as conn:
        print("Obteniendo todos los materiales MAT...")
        todos_mat = conn.execute(text(
            'SELECT "CodMat", "Descri", "CosMat" '
            'FROM cost360_materials '
            "WHERE \"CodMat\" LIKE 'MAT%' "
            'ORDER BY "CodMat"'
        )).fetchall()
        print(f"Total MAT a analizar: {len(todos_mat)}")

        print("Obteniendo materiales correctos (no-MAT)...")
        todos_no_mat = conn.execute(text(
            'SELECT "CodMat", "Descri", "CosMat" '
            'FROM cost360_materials '
            "WHERE \"CodMat\" NOT LIKE 'MAT%' "
            'ORDER BY "CodMat"'
        )).fetchall()
        print(f"Total materiales correctos: {len(todos_no_mat)}")

        # Índice de materiales no-MAT
        materiales_dict: dict = {}
        keywords_index: dict = {}

        for cod, descri, precio in todos_no_mat:
            descri_norm = normalizar_descripcion(descri)
            keywords    = extraer_palabras_clave(descri)
            diametro    = extraer_diametro(descri)

            materiales_dict[cod] = {
                'descri':      descri,
                'precio':      precio,
                'descri_norm': descri_norm,
                'keywords':    keywords,
                'diametro':    diametro,
            }
            for kw in keywords:
                keywords_index.setdefault(kw, []).append(cod)

        row_num = 2
        resultados = {'match_exacto': 0, 'match_parcial': 0, 'sin_match': 0, 'bloq_diam': 0}

        print("Procesando coincidencias (con validacion de diametro)...")

        for mat_cod, mat_descri, mat_precio in todos_mat:
            mat_norm     = normalizar_descripcion(mat_descri)
            mat_keywords = extraer_palabras_clave(mat_descri)
            mat_diam     = extraer_diametro(mat_descri)

            # ── Buscar match exacto (normalizado + diámetro compatible) ─────
            match_exacto = None
            for cod, datos in materiales_dict.items():
                if (datos['descri_norm'] == mat_norm
                        and diametros_compatibles(mat_diam, datos['diametro'])):
                    match_exacto = (cod, datos['descri'], datos['precio'], datos['diametro'])
                    break

            # ── Buscar por keywords ──────────────────────────────────────────
            match_keywords = None
            if not match_exacto and len(mat_keywords) >= 2:
                candidatos = set()
                for kw in mat_keywords:
                    if kw in keywords_index:
                        candidatos.update(keywords_index[kw])

                mejor = 0
                for cod in candidatos:
                    datos = materiales_dict[cod]
                    if not diametros_compatibles(mat_diam, datos['diametro']):
                        continue  # Diámetro incompatible → ignorar
                    kw_match = sum(1 for kw in mat_keywords if kw in datos['keywords'])
                    if kw_match > mejor and kw_match >= len(mat_keywords) * 0.7:
                        match_keywords = (cod, datos['descri'], datos['precio'], datos['diametro'], kw_match)
                        mejor = kw_match

            # ── Buscar fuzzy ─────────────────────────────────────────────────
            match_fuzzy = None
            mejor_score = 0
            if not match_exacto and not match_keywords and len(mat_keywords) >= 1:
                candidatos = set()
                for kw in mat_keywords:
                    if kw in keywords_index:
                        candidatos.update(keywords_index[kw])
                if not candidatos:
                    import random
                    candidatos = set(random.sample(list(materiales_dict.keys()),
                                                   min(100, len(materiales_dict))))
                for cod in candidatos:
                    datos = materiales_dict[cod]
                    if not diametros_compatibles(mat_diam, datos['diametro']):
                        continue  # Diámetro incompatible → ignorar
                    score = calcular_similitud(mat_norm, datos['descri_norm'])
                    if score > 0.85 and score > mejor_score:
                        match_fuzzy = (cod, datos['descri'], datos['precio'], datos['diametro'], score)
                        mejor_score = score

            # ── Determinar resultado ─────────────────────────────────────────
            if match_exacto:
                tipo_match   = "EXACTO"
                match_cod, match_descri, match_precio, match_diam = match_exacto
                resultados['match_exacto'] += 1
                fill         = match_exacto_fill
                propuesta    = "FUSIONAR"
                cod_propuesto = match_cod
                obs          = "Match exacto (normalizado)"
                score_value  = "100%"

            elif match_fuzzy:
                tipo_match   = "FUZZY"
                match_cod, match_descri, match_precio, match_diam, score = match_fuzzy
                resultados['match_parcial'] += 1
                fill         = match_parcial_fill
                propuesta    = "REVISAR" if score < 0.9 else "FUSIONAR"
                cod_propuesto = match_cod
                obs          = f"Fuzzy {score:.1%}"
                score_value  = f"{score:.1%}"

            elif match_keywords:
                tipo_match   = "KEYWORDS"
                match_cod, match_descri, match_precio, match_diam, kw_n = match_keywords
                resultados['match_parcial'] += 1
                fill         = match_parcial_fill
                propuesta    = "REVISAR"
                cod_propuesto = match_cod
                obs          = f"Keywords ({kw_n}/{len(mat_keywords)})"
                score_value  = f"{kw_n}/{len(mat_keywords)}"

            else:
                tipo_match    = "SIN MATCH"
                match_cod = match_descri = cod_propuesto = obs = ""
                match_precio  = 0
                match_diam    = None
                resultados['sin_match'] += 1
                fill          = sin_match_fill
                propuesta     = "MANUAL"
                score_value   = ""

            # Diferencias de precio
            diff_dolares = abs(mat_precio - match_precio) if match_precio else 0
            diff_pct = (
                (match_precio - mat_precio) / mat_precio * 100
                if mat_precio and match_precio else 0
            )

            row_data = [
                mat_cod,
                mat_descri,
                round(mat_precio, 2),
                mat_diam or "",
                "SI" if match_cod else "NO",
                match_cod,
                match_descri,
                round(match_precio, 2) if match_precio else "",
                match_diam or "",
                score_value,
                round(diff_dolares, 2) if diff_dolares else "",
                f"{diff_pct:.1f}%" if diff_pct else "",
                tipo_match,
                propuesta,
                cod_propuesto,
                obs,
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                if col_num in [3, 8, 11]:
                    if value and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'

            row_num += 1

            if row_num % 100 == 0:
                print(f"Procesados {row_num - 1} de {len(todos_mat)} registros...")

    # ── Resumen ──────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("RESUMEN DEL ANALISIS (con validacion de diametro)")
    print("=" * 60)
    total = len(todos_mat)
    for k, v in resultados.items():
        print(f"  {k:<15}: {v:>5}  ({v/total*100:.1f}%)")

    filename = f"analisis_mat_v2_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"\nArchivo generado: {filename}")


if __name__ == '__main__':
    main()
